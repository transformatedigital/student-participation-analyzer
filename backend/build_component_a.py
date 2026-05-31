#!/usr/bin/env python3
"""
Construye Componente A para una clase:
  - JSON estructurado (para el dashboard de la plataforma)
  - Excel con DOS secciones claras (Preparation / Participation)
    siguiendo la plantilla de la maestra para que pueda enlazar fórmulas.

Reglas de scoring (escala 1–6):
  PREPARATION (7.5%) — Questions to students
    · Pregunta dirigida + respondió → 1–6 (calidad de la respuesta)
    · Pregunta dirigida + NO respondió → 0 (penaliza falta de preparación)
    · Pregunta ABIERTA al grupo + respondió → 1–6
    · Pregunta abierta + no respondió → N/A
    · Pregunta dirigida a otro alumno → N/A

  PARTICIPATION (7.5%) — Questions to Professor + voluntary
    · Pregunta del alumno al profe → 1–6
    · Comentario voluntario del alumno → 1–6
    · Otros alumnos en esa fila → N/A

  ATTENDANCE (5%): manual, editable en attendance.json

Mapeo de evaluaciones existentes A–E → 1–6:
    A→6   B→5   C→4   D→2   E→1
"""

import json
import os
from pathlib import Path
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Funciones de scoring con la rúbrica oficial (1-6) — leen el texto directamente
import sys
sys.path.insert(0, str(Path(__file__).parent))
from score_with_rubric import score_response, score_question, score_comment

REPO_ROOT = Path(__file__).resolve().parent.parent
SESSION_ID = os.getenv("SESSION_ID", "2026-04-07")
SESSION_DIR = REPO_ROOT / "data" / "clases" / SESSION_ID
ANALYSIS_FILE = SESSION_DIR / "analysis.json"
CACHE_DIR = SESSION_DIR / "transcript_cache"
EVALS_FILE = SESSION_DIR / "ai_evaluations.json"
ATTENDANCE_FILE = SESSION_DIR / "attendance.json"
OUTPUT_JSON = SESSION_DIR / "component_a.json"
OUTPUT_XLSX = SESSION_DIR / f"Component_A_{SESSION_ID}.xlsx"

STUDENTS = ["Aryang", "Mega", "Chilaka", "Grace", "Sthepen"]
# Mapeo CONSERVADOR: el 6 se reserva para respuestas excepcionales
# (que la heurística no puede detectar). Por defecto A→5, B→4, C→3.
# Refinamiento manual posterior puede subir a 6 los casos meritorios.
GRADE_TO_1_6 = {"A": 5, "B": 4, "C": 3, "D": 2, "E": 1}
PREP_WEIGHT, PART_WEIGHT, ATT_WEIGHT = 7.5, 7.5, 5.0
STUDENT_TYPES = {"student_response", "student_question", "student_comment"}
EXCLUDED_SPEAKERS = {"UNKNOWN", "ILEANG", "ILEANA", ""}


def load_attendance():
    if ATTENDANCE_FILE.exists():
        with open(ATTENDANCE_FILE) as f:
            return json.load(f)
    return {s: True for s in STUDENTS}


def load_evals():
    if not EVALS_FILE.exists():
        return {}
    with open(EVALS_FILE, encoding="utf-8") as f:
        data = json.load(f)
    return {e["id"]: e["grade"] for e in data.get("evaluations", [])}


# Scores manuales 1–6 generados con la rúbrica oficial (override del mapeo A–E)
MANUAL_SCORES_FILE = SESSION_DIR / "manual_scores_1_6.json"
def load_manual_scores():
    if MANUAL_SCORES_FILE.exists():
        with open(MANUAL_SCORES_FILE, encoding="utf-8") as f:
            return {int(k): int(v) for k, v in json.load(f).items()}
    return {}


# Reconstruir IDs en el mismo orden que evaluate_with_ai.py para mapear scores
def build_participation_index():
    parts = []
    for f in sorted(CACHE_DIR.glob("block_*.json")):
        with open(f, encoding="utf-8") as fh:
            block = json.load(fh)
        for u in block.get("utterances", []):
            sp = u.get("speaker", "")
            if u.get("type") not in STUDENT_TYPES or sp in EXCLUDED_SPEAKERS:
                continue
            parts.append({
                "id": len(parts) + 1,
                "block": block["block"],
                "speaker": sp.title(),  # ARYANG -> Aryang
                "type": u["type"],
                "text": u.get("text", ""),
                "timestamp": u.get("timestamp_abs") or u.get("timestamp", ""),
            })
    return parts


# Carga datos ─────────────────────────────────────────────────────────────────
with open(ANALYSIS_FILE, encoding="utf-8") as f:
    analysis = json.load(f)

evals = load_evals()
manual_scores = load_manual_scores()  # id → score 1-6 (rúbrica oficial)
parts_index = build_participation_index()
attendance = load_attendance()

USING_MANUAL = bool(manual_scores)

# Index participaciones por (speaker, type, text) para encontrar score
parts_lookup = {}
parts_lookup_id = {}
for p in parts_index:
    key = (p["speaker"].lower(), p["type"], p["text"][:80].lower())
    parts_lookup[key] = p["id"]
    parts_lookup_id[p["id"]] = (p["speaker"], p["type"], p["text"])


def find_score(speaker, ptype, text):
    """Devuelve score 1-6 priorizando scores manuales sobre mapeo A-E."""
    if not speaker or not text:
        return None
    key = (speaker.lower(), ptype, text[:80].lower())
    pid = parts_lookup.get(key)
    if pid and pid in manual_scores:
        return manual_scores[pid]
    g = evals.get(pid) if pid else None
    return GRADE_TO_1_6.get(g) if g else None


# ─── PREPARATION: filas = preguntas del profe ───────────────────────────────
prep_rows = []
for q in analysis.get("teacher_questions", []):
    directed_to = (q.get("directed_to") or "").strip()
    is_open = directed_to.lower() in ("class", "everyone", "all", "open", "")
    sr = q.get("student_response") or {}
    responder = (sr.get("student") or "").strip()
    response_text = sr.get("response", "")
    response_grade = sr.get("quality_score", "")

    row = {
        "timestamp": q.get("timestamp", ""),
        "block": q.get("block"),
        "question": q.get("question", "").strip(),
        "directed_to": directed_to if not is_open else "ABIERTA",
        "is_open": is_open,
        "responder": responder.title() if responder else None,
        "response_text": response_text,
        "scores": {s: None for s in STUDENTS},  # None = N/A
    }

    # Asignar scores según reglas de Preparation
    if not is_open and directed_to:
        # Dirigida a alguien específico
        target = directed_to.title()
        if target in STUDENTS:
            if responder and responder.title() == target:
                # Respondió → score 1–6 con la rúbrica oficial directamente
                score = score_response(response_text, response_grade)
                row["scores"][target] = score
            else:
                # No respondió la pregunta dirigida → 0
                row["scores"][target] = 0
        # Otros alumnos: N/A (siguen None)
    else:
        # Abierta: solo cuenta para quien respondió
        if responder:
            r = responder.title()
            if r in STUDENTS:
                score = score_response(response_text, response_grade)
                row["scores"][r] = score

    prep_rows.append(row)


# ─── PARTICIPATION: filas = contribuciones voluntarias ──────────────────────
# Fuente: transcript_cache + ai_evaluations (cubre los 4 alumnos, no solo Aryang/Grace)
part_rows = []
TYPE_LABEL = {
    "student_question": "Question to Professor",
    "student_comment": "Comment",
    "student_response": "Voluntary Response",
}
for p in parts_index:
    student = p["speaker"]
    if student not in STUDENTS:
        continue
    ctype = p["type"]
    # student_response cuando NO es respuesta directa a una pregunta dirigida
    # (las dirigidas ya entran en Preparation arriba). Para simplificar:
    # student_question y student_comment siempre van a Participation.
    # student_response va a Participation SOLO si no la consumió Preparation.
    if ctype == "student_response":
        # Si esta respuesta ya está mapeada a una pregunta dirigida en Preparation,
        # no la duplicar. Heurística: si fue "dirigida" a este alumno, ya cuenta arriba.
        already_in_prep = any(
            r.get("responder") == student and r["scores"].get(student) is not None
            and r["response_text"][:50].lower() in p["text"].lower()
            for r in prep_rows if not r.get("is_open")
        )
        if already_in_prep:
            continue
    if ctype not in TYPE_LABEL:
        continue

    # Prioriza scores manuales (rúbrica 1-6 oficial); fallback al mapeo A-E
    score = manual_scores.get(p["id"])
    if not score:
        grade = evals.get(p["id"])
        score = GRADE_TO_1_6.get(grade) if grade else None
    if not score:
        score = 3  # default conservador

    row = {
        "timestamp": p["timestamp"],
        "block": p["block"],
        "speaker": student,
        "type": TYPE_LABEL[ctype],
        "text": p["text"],
        "scores": {s: None for s in STUDENTS},
    }
    row["scores"][student] = score
    part_rows.append(row)

# Ordenar por timestamp
def ts_to_seconds(ts):
    try:
        parts = str(ts).split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return 0

part_rows.sort(key=lambda r: ts_to_seconds(r["timestamp"]))


# ─── Cálculos por alumno ────────────────────────────────────────────────────
def aggregate(rows, weight):
    """Devuelve dict: student → (sum_scores, n_opportunities, max_possible, pct_grade)"""
    out = {}
    for s in STUDENTS:
        total = 0
        n_opp = 0
        for r in rows:
            v = r["scores"].get(s)
            if v is None:
                continue  # N/A no cuenta como oportunidad
            total += v
            n_opp += 1
        max_possible = n_opp * 6
        pct = (total / max_possible * weight) if max_possible > 0 else 0
        out[s] = {
            "sum_scores": total,
            "n_opportunities": n_opp,
            "max_possible": max_possible,
            "pct_obtained": round(pct, 2),
            "pct_max": weight,
        }
    return out


prep_totals = aggregate(prep_rows, PREP_WEIGHT)
part_totals = aggregate(part_rows, PART_WEIGHT)

# Resumen final — DOS ESCENARIOS
# Escenario A: literal del silabo, solo calidad promedio (lo de ahora)
# Escenario B: misma calidad + bonus discreto por cantidad de participaciones
#              Bonus en Participation: 0=+0  1-4=+0  5-14=+0.5  15-29=+1.0  30+=+1.5
#              (capped a 7.5 que es el máximo del componente)
def quantity_bonus(n_contribs):
    if n_contribs >= 30:
        return 1.5
    if n_contribs >= 15:
        return 1.0
    if n_contribs >= 5:
        return 0.5
    return 0.0


summary = {}
for s in STUDENTS:
    p = prep_totals[s]["pct_obtained"]
    pa_quality = part_totals[s]["pct_obtained"]
    n_part = part_totals[s]["n_opportunities"]
    bonus = quantity_bonus(n_part)
    pa_with_bonus = min(PART_WEIGHT, round(pa_quality + bonus, 2))

    asistio = attendance.get(s, True)
    att = ATT_WEIGHT if asistio else 0.0

    total_a_scenario_a = round(p + pa_quality + att, 2)
    total_a_scenario_b = round(p + pa_with_bonus + att, 2)

    summary[s] = {
        "preparation_pct": p,
        "participation_pct_A_quality_only": pa_quality,
        "participation_pct_B_with_bonus": pa_with_bonus,
        "quantity_bonus": bonus,
        "n_voluntary_contribs": n_part,
        "attendance_pct": att,
        "attendance_status": "Asistió" if asistio else "No asistió",
        "component_a_total_scenario_A": total_a_scenario_a,
        "component_a_total_scenario_B": total_a_scenario_b,
        "component_a_max": 20.0,
    }

# ─── Guardar JSON ───────────────────────────────────────────────────────────
output = {
    "session_id": SESSION_ID,
    "generated_at": datetime.now().isoformat(timespec="seconds"),
    "students": STUDENTS,
    "attendance": attendance,
    "rubric": {
        "scale": "1–6 (mapeado de evaluaciones A–E: A→6, B→5, C→4, D→2, E→1)",
        "weights": {"preparation": PREP_WEIGHT, "participation": PART_WEIGHT, "attendance": ATT_WEIGHT},
    },
    "preparation": {
        "rows": prep_rows,
        "totals": prep_totals,
    },
    "participation": {
        "rows": part_rows,
        "totals": part_totals,
    },
    "summary": summary,
}

with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2, ensure_ascii=False)


# ─── Generar Excel ──────────────────────────────────────────────────────────
wb = Workbook()

# Estilos
HEADER_FILL = PatternFill("solid", fgColor="305496")
SUBHEADER_FILL = PatternFill("solid", fgColor="8EA9DB")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
PCT_FILL = PatternFill("solid", fgColor="C6EFCE")
NA_FILL = PatternFill("solid", fgColor="F2F2F2")
ZERO_FILL = PatternFill("solid", fgColor="FFC7CE")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_section(ws, title, subtitle, rows, totals, weight, start_row, item_label):
    """Escribe una sección Preparation o Participation."""
    r = start_row
    # Título
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(STUDENTS) + 2)
    r += 1

    ws.cell(row=r, column=1, value=subtitle).font = Font(italic=True, size=10, color="595959")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(STUDENTS) + 2)
    r += 2

    # Encabezado de tabla
    headers = ["Time", item_label] + STUDENTS + ["Notes / Feedback"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = WHITE
        c.fill = SUBHEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    r += 1

    # Filas
    for row_data in rows:
        ws.cell(row=r, column=1, value=row_data.get("timestamp", "")).alignment = CENTER
        ws.cell(row=r, column=1).border = BORDER

        text_col = row_data.get("question") or row_data.get("text", "")
        if "directed_to" in row_data:
            tag = f" → {row_data['directed_to']}" if not row_data.get("is_open") else " (open)"
            text_col = text_col + tag
        if "type" in row_data and "directed_to" not in row_data:
            text_col = f"[{row_data['type']}] {text_col}"

        c = ws.cell(row=r, column=2, value=text_col)
        c.alignment = LEFT
        c.border = BORDER

        for col, s in enumerate(STUDENTS, start=3):
            v = row_data["scores"].get(s)
            if v is None:
                cell = ws.cell(row=r, column=col, value="")
                cell.fill = NA_FILL
            elif v == 0:
                cell = ws.cell(row=r, column=col, value=0)
                cell.fill = ZERO_FILL
            else:
                cell = ws.cell(row=r, column=col, value=v)
            cell.alignment = CENTER
            cell.border = BORDER

        ws.cell(row=r, column=3 + len(STUDENTS), value="").border = BORDER
        r += 1

    # Fila TOTAL
    ws.cell(row=r, column=2, value="TOTAL (sum of scores)").font = BOLD
    ws.cell(row=r, column=2).fill = TOTAL_FILL
    ws.cell(row=r, column=2).border = BORDER
    for col, s in enumerate(STUDENTS, start=3):
        c = ws.cell(row=r, column=col, value=totals[s]["sum_scores"])
        c.font = BOLD
        c.fill = TOTAL_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.cell(row=r, column=3 + len(STUDENTS)).fill = TOTAL_FILL
    ws.cell(row=r, column=3 + len(STUDENTS)).border = BORDER
    r += 1

    # Fila Max possible
    ws.cell(row=r, column=2, value="Max possible (n_opp × 6)").font = Font(italic=True, color="595959")
    ws.cell(row=r, column=2).border = BORDER
    for col, s in enumerate(STUDENTS, start=3):
        c = ws.cell(row=r, column=col, value=totals[s]["max_possible"])
        c.font = Font(italic=True, color="595959")
        c.alignment = CENTER
        c.border = BORDER
    ws.cell(row=r, column=3 + len(STUDENTS)).border = BORDER
    r += 1

    # Fila % de weight
    ws.cell(row=r, column=2, value=f"% Grade (max {weight}%)").font = BOLD
    ws.cell(row=r, column=2).fill = PCT_FILL
    ws.cell(row=r, column=2).border = BORDER
    for col, s in enumerate(STUDENTS, start=3):
        c = ws.cell(row=r, column=col, value=totals[s]["pct_obtained"])
        c.font = BOLD
        c.fill = PCT_FILL
        c.number_format = "0.00"
        c.alignment = CENTER
        c.border = BORDER
    ws.cell(row=r, column=3 + len(STUDENTS)).fill = PCT_FILL
    ws.cell(row=r, column=3 + len(STUDENTS)).border = BORDER
    return r + 2


# Hoja 1: Componente A — clase única
ws1 = wb.active
ws1.title = "Component A"

ws1.cell(row=1, column=1, value=f"COMPONENTE A — Class {SESSION_ID}").font = Font(bold=True, size=16, color="305496")
ws1.cell(row=2, column=1, value="Student's Preparation, Participation & Attendance (20% del total)").font = Font(italic=True, size=11, color="595959")

# Asistencia
ws1.cell(row=4, column=1, value="ATTENDANCE (5%) — editable").font = Font(bold=True, size=12, color="FFFFFF")
ws1.cell(row=4, column=1).fill = HEADER_FILL
ws1.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
for i, s in enumerate(STUDENTS):
    ws1.cell(row=5, column=2 + i, value=s).font = WHITE
    ws1.cell(row=5, column=2 + i).fill = SUBHEADER_FILL
    ws1.cell(row=5, column=2 + i).alignment = CENTER
    ws1.cell(row=5, column=2 + i).border = BORDER
    asistio = attendance.get(s, True)
    val = "Asistió" if asistio else "No asistió"
    c = ws1.cell(row=6, column=2 + i, value=val)
    c.alignment = CENTER
    c.border = BORDER
    if asistio:
        c.fill = PCT_FILL
    else:
        c.fill = ZERO_FILL
ws1.cell(row=6, column=1, value="Status (manual)").font = Font(italic=True)

# Sección Preparation
next_row = 9
next_row = write_section(
    ws1, "PREPARATION (7.5%) — Questions from Professor",
    "Pregunta dirigida + respondió: 1–6 · No respondió: 0 · Pregunta abierta + respondió: 1–6 · Otros: N/A",
    prep_rows, prep_totals, PREP_WEIGHT, next_row, "Question"
)

# Sección Participation
next_row = write_section(
    ws1, "PARTICIPATION (7.5%) — Questions from Student / Voluntary",
    "Pregunta del alumno o comentario voluntario: 1–6 · Otros alumnos en esa fila: N/A",
    part_rows, part_totals, PART_WEIGHT, next_row, "Contribution"
)

# Resumen final — COMPARATIVA DOS ESCENARIOS
ws1.cell(row=next_row, column=1, value="RESUMEN COMPONENTE A — COMPARATIVA DOS ESCENARIOS").font = Font(bold=True, size=14, color="FFFFFF")
ws1.cell(row=next_row, column=1).fill = HEADER_FILL
ws1.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=10)
next_row += 1

ws1.cell(row=next_row, column=1,
    value="Escenario A: literal del silabo (solo calidad promedio). Escenario B: igual + bonus por cantidad de contribuciones (5-14:+0.5  15-29:+1.0  30+:+1.5)").font = Font(italic=True, size=10, color="595959")
ws1.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=10)
next_row += 2

resume_headers = ["Student", "Prep (7.5)", "Att (5)", "─ Esc A ─", "Part A", "TOTAL A", "% A", "─ Esc B ─", "Part B (+bonus)", "TOTAL B", "% B"]
for col, h in enumerate(resume_headers, start=1):
    c = ws1.cell(row=next_row, column=col, value=h)
    c.font = WHITE; c.fill = SUBHEADER_FILL; c.alignment = CENTER; c.border = BORDER
next_row += 1

for s in STUDENTS:
    sm = summary[s]
    pct_a = round(sm["component_a_total_scenario_A"] / 20 * 100, 1)
    pct_b = round(sm["component_a_total_scenario_B"] / 20 * 100, 1)
    bonus_label = f"+{sm['quantity_bonus']}" if sm['quantity_bonus'] else "—"
    part_b_label = f"{sm['participation_pct_B_with_bonus']} ({bonus_label})"
    vals = [
        s,
        sm["preparation_pct"],
        sm["attendance_pct"],
        "",  # separador
        sm["participation_pct_A_quality_only"],
        sm["component_a_total_scenario_A"],
        pct_a,
        "",  # separador
        part_b_label,
        sm["component_a_total_scenario_B"],
        pct_b,
    ]
    for col, v in enumerate(vals, start=1):
        c = ws1.cell(row=next_row, column=col, value=v)
        c.alignment = CENTER; c.border = BORDER
        if col == 1:
            c.font = BOLD
        if col == 6:
            c.fill = PatternFill("solid", fgColor="FFF2CC"); c.font = BOLD
            c.number_format = "0.00"
        if col == 10:
            c.fill = PatternFill("solid", fgColor="C6EFCE"); c.font = BOLD
            c.number_format = "0.00"
        if col in (2, 3, 5):
            c.number_format = "0.00"
        if col in (7, 11):
            c.number_format = "0.0\"%\""
    next_row += 1

next_row += 1
ws1.cell(row=next_row, column=1, value="🟡 Escenario A — apego literal al silabo").font = Font(bold=True, color="996600")
next_row += 1
ws1.cell(row=next_row, column=1, value="🟢 Escenario B — sugerido: reconoce engagement sin romper rúbrica").font = Font(bold=True, color="00691E")
next_row += 1

# Hoja 2: Rúbrica de referencia
ws2 = wb.create_sheet("Rúbrica 1–6")
ws2.cell(row=1, column=1, value="Rúbrica de Class Preparation & Participation (escala 1–6)").font = Font(bold=True, size=14, color="305496")
rubric = [
    (6, "Comes prepared. Contributes readily without dominating. Thoughtful contributions advance the conversation. Respects others' views. Active in small groups."),
    (5, "Comes prepared, makes thoughtful comments when called. Contributes occasionally without prompting. Respects others. Active in small groups."),
    (4, "Generally prepared. Participates but may dominate, ramble, interrupt with digressive questions, bluff when unprepared, miss social cues."),
    (3, "Comes prepared. Does NOT voluntarily contribute. Minimal answers when called. Listens attentively, takes notes."),
    (2, "Comes but NOT prepared. Does not contribute. Unable to contribute usefully when called."),
    (1, "Comes but NOT prepared. May be disruptive. Negative impact on others."),
]
ws2.cell(row=3, column=1, value="Score").font = WHITE; ws2.cell(row=3, column=1).fill = SUBHEADER_FILL
ws2.cell(row=3, column=2, value="Descriptor").font = WHITE; ws2.cell(row=3, column=2).fill = SUBHEADER_FILL
for i, (lvl, desc) in enumerate(rubric, start=4):
    ws2.cell(row=i, column=1, value=lvl).font = BOLD
    ws2.cell(row=i, column=1).alignment = CENTER
    ws2.cell(row=i, column=2, value=desc).alignment = LEFT
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 100

# Anchos de columna en hoja 1
ws1.column_dimensions["A"].width = 10
ws1.column_dimensions["B"].width = 60
for i, _ in enumerate(STUDENTS, start=3):
    ws1.column_dimensions[get_column_letter(i)].width = 12
ws1.column_dimensions[get_column_letter(3 + len(STUDENTS))].width = 30

wb.save(OUTPUT_XLSX)

# ─── Reporte por consola ────────────────────────────────────────────────────
print("=" * 78)
print(f"  COMPONENTE A — Clase {SESSION_ID}")
print("=" * 78)
print(f"\nPREPARATION (7.5%) — {len(prep_rows)} preguntas del profe")
print(f"  · Dirigidas: {sum(1 for r in prep_rows if not r['is_open'])}")
print(f"  · Abiertas:  {sum(1 for r in prep_rows if r['is_open'])}")
for s in STUDENTS:
    t = prep_totals[s]
    print(f"  {s:10s}: {t['sum_scores']:>3} pts / {t['max_possible']:>3} max  → {t['pct_obtained']:>5.2f} / 7.5  ({t['n_opportunities']} oportunidades)")

print(f"\nPARTICIPATION (7.5%) — {len(part_rows)} contribuciones voluntarias")
for s in STUDENTS:
    t = part_totals[s]
    print(f"  {s:10s}: {t['sum_scores']:>3} pts / {t['max_possible']:>3} max  → {t['pct_obtained']:>5.2f} / 7.5  ({t['n_opportunities']} contribs)")

print(f"\nATTENDANCE (5%):")
for s in STUDENTS:
    print(f"  {s:10s}: {summary[s]['attendance_status']:12s} → {summary[s]['attendance_pct']:.1f} / 5.0")

print(f"\n{'═'*78}")
print(f"  COMPARATIVA DOS ESCENARIOS — Componente A (sobre 20)")
print(f"{'═'*78}")
print(f"\n  Escenario A: literal del silabo (solo calidad promedio)")
print(f"  Escenario B: igual + bonus por cantidad (5-14:+0.5  15-29:+1.0  30+:+1.5)\n")

print(f"  {'Alumno':10s}  {'Prep':>5s}  {'Att':>4s}  ║ {'Part A':>6s}  {'TOTAL A':>8s}  {'% A':>5s}  ║ {'Part B':>10s}  {'TOTAL B':>8s}  {'% B':>5s}")
print(f"  {'-'*10}  {'-'*5}  {'-'*4}  ║ {'-'*6}  {'-'*8}  {'-'*5}  ║ {'-'*10}  {'-'*8}  {'-'*5}")
for s in STUDENTS:
    sm = summary[s]
    pct_a = sm["component_a_total_scenario_A"] / 20 * 100
    pct_b = sm["component_a_total_scenario_B"] / 20 * 100
    bonus_str = f"+{sm['quantity_bonus']}" if sm['quantity_bonus'] else ""
    part_b_str = f"{sm['participation_pct_B_with_bonus']:>4.2f} {bonus_str:>4s}"
    print(f"  {s:10s}  {sm['preparation_pct']:>5.2f}  {sm['attendance_pct']:>4.1f}  ║ {sm['participation_pct_A_quality_only']:>6.2f}  {sm['component_a_total_scenario_A']:>8.2f}  {pct_a:>4.1f}%  ║ {part_b_str:>10s}  {sm['component_a_total_scenario_B']:>8.2f}  {pct_b:>4.1f}%")

print(f"\n✅ JSON  → {OUTPUT_JSON.relative_to(REPO_ROOT)}")
print(f"✅ Excel → {OUTPUT_XLSX.relative_to(REPO_ROOT)}")
print(f"   (Edita asistencia en {ATTENDANCE_FILE.relative_to(REPO_ROOT)} y vuelve a correr)")
