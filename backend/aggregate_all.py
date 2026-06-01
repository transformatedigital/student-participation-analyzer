#!/usr/bin/env python3
"""
Agrega Componente A de todas las clases procesadas y genera:

  1. data/aggregated/all_classes.json — JSON con resumen por alumno y por clase
  2. data/aggregated/Component_A_Aggregated.xlsx — Excel acumulado para la maestra
  3. data/aggregated/dashboard.html — Dashboard standalone (sin servidor)

Lee cada data/clases/<sid>/component_a.json y promedia/suma por alumno.
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data" / "clases"
OUT_DIR = REPO_ROOT / "docs"  # GitHub Pages reconoce /docs/ automáticamente
OUT_DIR.mkdir(parents=True, exist_ok=True)

STUDENTS = ["Aryang", "Mega", "Chilaka", "Grace", "Sthepen"]

# Mapping email → student name (para Componente B cross-evaluations)
EMAIL_TO_STUDENT = {
    "megawatisuharsonoputri@gmail.com": "Mega",
    "aryacgs@gmail.com": "Aryang",
    "gakubastephen@gmail.com": "Sthepen",
    "chijioke.chilaka@gmail.com": "Chilaka",
    "songanzilagracia333@gmail.com": "Grace",
}
# Mapping nombres alternativos del presenter → student name canónico
PRESENTER_ALIASES = {
    "Chilaka": "Chilaka", "Chillaka": "Chilaka", "CHILAKA": "Chilaka",
    "Grace": "Grace", "GRACE": "Grace",
    "Mega": "Mega", "MEGA": "Mega", "Megawat": "Mega", "Megawati": "Mega",
    "Aryang": "Aryang", "Arya": "Aryang", "ARYA": "Aryang",
    "Arya Cipta Graha Sinaga": "Aryang",
    "Sthepen": "Sthepen", "Stephen": "Sthepen", "STEPHEN": "Sthepen",
    "Gakuba Stephen": "Sthepen", "Gakuba": "Sthepen",
}


def get_component_c_data():
    """C.1 Assignments — 5 exams during semester. Sum / max = % of 10%."""
    exam1_responses = {
        "Grace": {
            "full_name": "SONGANZILA GRACE MFUMU",
            "answers": [
                "The goal of this class is to permit students or researchers to understand how the technology is implemented from the lab to the market following different steps: identification of the problem, conceptualization, implementation and commercialization. But specifically in the context of GDI, this course is much more important allowing students to build a strong partnership between Korea and their home countries through technology knowledge transfer from a Korean company to solve an identified technology problem.",
                "The real problem is about improving the effectiveness of the academic system in DRC. Universities in DRC operate manually without a centralized system for academic management. As a result, college students are left behind from digital practice.",
                "The owner in this case is the University of Kinshasa.",
                "(No answer provided)",
            ],
            "scores": [25, 25, 25, 25],
        },
        "Sthepen": {
            "full_name": "GAKUNBA STEPHENS",
            "answers": [
                "This course is mainly focusing on two things: (1) How to solve problems that are related with IT in the world — problems are in two types: Problem driven (problem comes from citizens to the lab) and Technology driven (technology from lab to market). In simple terms: GTC is the process where we take IT innovations from the lab to market. This course will help us in our ICP to understand if it is problem driven or technology push.",
                "The main problem in my country is the 'Real time gap' monitoring for the frauds happening on mobile phone via SMS and voice communications. In Rwanda, mobile money is used for transfers and buying goods. Although SIM and ID are registered by biometric verification, we still need to know how the SIM card is used after registration to be proactive and protect citizens from losing money in social engineering activities.",
                "This is the problem driven case but the main owner is RURA and NIDA: as government agencies we need to protect our citizens through e-services and e-payment.",
                "To have a secure e-payment. To have e-services. (Response incomplete)",
            ],
            "scores": [25, 25, 25, 25],
        },
        "Chilaka": {
            "full_name": "CHILAKA CHIJIOKE REGINALD",
            "answers": [
                "To better understand global technology commercialization in the context of our countries. To understand how technology transfer can help solve the most important needs of our country. To also better understand a very good logic flow towards accomplishing technology transfer and finally to better understand our roles as intermediaries — as a bridge between our country and Korea.",
                "The urgent call by the government of my country on how to utilize research outcomes in the ecosystem to become commercializable products to help the economy gave rise to my country's current GTC initiative. It aligns perfectly with the government initiative called Energize Commercialization Now (ECON), which looks for a way to manage and monitor R&D outcomes for sustainable development. This can be better solved with the adaptation of the NTIS system of Korea.",
                "My organization, Federal Ministry of Innovation, Science and Technology will be the owner and will be domiciled in the National Board for Technology Incubation (NBTI) as it lies within their mandate.",
                "1) More products and services from universities/research agencies instead of ideas ending on the shelf. 2) R&D can easily be monitored through the process lifecycle. 3) Investments in R&D can easily be followed up to see the outcome. 4) Will help the economy by reducing imports of products/services produced within the country. 5) Funding for inventors and innovators from investors.",
            ],
            "scores": [25, 25, 25, 25],
        },
        "Aryang": {
            "full_name": "ARYA C.G.S",
            "answers": [
                "Commercialize the technology from lab to market. Like in my case, how to develop the TTS model from lab and sell it to market (B&I) to increase the value of that invention. Start from ICP to GRC and make it happen.",
                "1. I haven't met the final partner that wants to do collaboration with me to develop the TTS model for localization of the Indonesian voice AI. 2. Budget??? I haven't thought for that case.",
                "The owner is GDM Division (they handle AI). They will be the one that develops and operates this product. Now they also want to confirm the legal team approach with Totocos lab.",
                "1. TTS model with high accuracy for Indonesian. 2. Combine with BRI chatbot. 3. Maybe move to another dialect. 4. Combine with BRI data to make personalized AI.",
            ],
            "scores": [25, 25, 25, 25],
        },
        "Mega": {
            "full_name": "MEGA",
            "answers": [
                "Global Technology Commercialization is the way to commercialization from research or innovation to be invention (market & industry).",
                "The real problem in my country is decarbonization. Indonesia has a commitment in the Paris Agreement for net zero emission. President made Regulation No. 20, 2020 about decarbonization in Indonesia — phase out coal power plants and change to renewable energy. One energy source very potential in Indonesia is photovoltaic (PV) power plants. But PV only produces during morning/afternoon; when night and rain come, it can't produce electricity, so we need technology to make grid stability and save energy. PV is also the target of the President in Indonesia (March 2026) — every island and province must have a PV power plant.",
                "National Research and Innovation Agency (BRIN).",
                "Impact for Indonesia: - Feasibility study for grid stability (policy recommendation). - To get the best technology from Korea. Impact for Korea: - Knowing the market especially in PV power plant in Indonesia because it is a very potential market.",
            ],
            "scores": [25, 25, 25, 25],
        },
    }

    questions_e1 = [
        "Q1: Objective of GTC course",
        "Q2: Real problem of GTC initiative",
        "Q3: Owner of the initiative",
        "Q4: Expected impact",
    ]

    exams = [
        {
            "id": 1, "date": "2026-05-12", "date_label": "Tue, May 12 (Wk 11)",
            "title": "Exam 1 — GTC Course Foundations",
            "max_per_question": 25, "n_questions": 4,
            "questions": questions_e1,
            "responses": exam1_responses,
            "status": "completed",
        },
        {"id": 2, "date": "2026-05-19", "date_label": "Tue, May 19 (Wk 12)",
         "title": "Exam 2 — Pending", "status": "pending"},
        {"id": 3, "date": "2026-05-26", "date_label": "Tue, May 26 (Wk 13)",
         "title": "Exam 3 — Pending", "status": "pending"},
        {"id": 4, "date": "2026-06-02", "date_label": "Tue, Jun 2 (Wk 14)",
         "title": "Exam 4 — Pending", "status": "pending"},
        {"id": 5, "date": "2026-06-09", "date_label": "Tue, Jun 9 (Wk 15)",
         "title": "Exam 5 — Pending", "status": "pending"},
    ]
    return {"exams": exams, "weight_total": 10.0, "n_exams": 5}


def get_component_b_data():
    """Cross-evaluations data — Week 2 (Mar 10, 2026): First Draft ICP Presentations."""
    raw = [
        # timestamp, evaluator_email, c1, c2, c3, c4, c5, comment, presenter_raw
        ("13:41:16", "aryacgs@gmail.com", 6, 5, 6, 5, 5, "interesting idea", "Chilaka"),
        ("13:44:56", "gakubastephen@gmail.com", 6, 6, 6, 6, 6, "all was good", "Chilaka"),
        ("14:01:40", "aryacgs@gmail.com", 5, 5, 6, 6, 5, "interesting idea, split slides to highlight point", "Grace"),
        ("14:02:59", "gakubastephen@gmail.com", 6, 6, 6, 6, 6, "It was good", "Grace"),
        ("14:22:42", "aryacgs@gmail.com", 5, 5, 5, 5, 5, "interesting idea, more eye contact, add more how its distributed", "Mega"),
        ("14:26:12", "megawatisuharsonoputri@gmail.com", 6, 6, 6, 6, 6, "", "Chilaka"),
        ("14:39:04", "gakubastephen@gmail.com", 6, 6, 6, 6, 6, "GOOD", "MEGA"),
        ("14:40:18", "megawatisuharsonoputri@gmail.com", 6, 6, 6, 6, 6, "Grace has to make milestone", "Grace"),
        ("14:41:13", "chijioke.chilaka@gmail.com", 5, 5, 3, 6, 3, "Focused on the power situation", "Megawat"),
        ("14:45:17", "songanzilagracia333@gmail.com", 3, 4, 4, 6, 4, "", "Mega"),
        ("14:45:34", "megawatisuharsonoputri@gmail.com", 6, 6, 6, 6, 6, "great idea", "Arya Cipta Graha Sinaga"),
        ("14:47:12", "chijioke.chilaka@gmail.com", 4, 5, 3, 5, 4, "Focused on a problem", "Arya"),
        ("14:48:35", "songanzilagracia333@gmail.com", 3, 3, 3, 4, 4, "", "Chillaka"),
        ("14:55:31", "gakubastephen@gmail.com", 6, 6, 6, 6, 6, "GOOD", "ARYA"),
        ("14:56:50", "songanzilagracia333@gmail.com", 3, 4, 3, 4, 4, "Arya need more clarity about the real problem, whether problem driven or technology push", "Arya"),
        ("15:28:33", "megawatisuharsonoputri@gmail.com", 6, 6, 6, 6, 6, "Great idea", "Gakuba Stephen"),
        ("15:28:46", "aryacgs@gmail.com", 6, 5, 5, 5, 6, "interesting idea, slides was great", "Stephen"),
        ("15:31:47", "songanzilagracia333@gmail.com", 4, 4, 3, 5, 4, "It was specific and clear, but comple. So he needs to work hard to find specific problems and Technology adoption", "Stephen"),
    ]
    DIMS = ["clarity", "structure", "audience", "conciseness", "delivery"]
    evals = []
    for ts, email, c1, c2, c3, c4, c5, comment, presenter_raw in raw:
        evaluator = EMAIL_TO_STUDENT.get(email, email)
        presenter = PRESENTER_ALIASES.get(presenter_raw.strip(), presenter_raw)
        if presenter not in STUDENTS:
            continue
        evals.append({
            "timestamp": ts, "evaluator": evaluator, "presenter": presenter,
            "clarity": c1, "structure": c2, "audience": c3, "conciseness": c4, "delivery": c5,
            "total": c1 + c2 + c3 + c4 + c5, "comment": comment,
        })

    # Agrupar por presenter
    per_presenter = {s: [] for s in STUDENTS}
    for e in evals:
        per_presenter[e["presenter"]].append(e)

    summary = {}
    for s, evs in per_presenter.items():
        if not evs:
            summary[s] = {"n_evals": 0, "avg_total": 0, "pct_b1": 0,
                          "by_dim": {d: 0 for d in DIMS}}
            continue
        avg_total = sum(e["total"] for e in evs) / len(evs)
        by_dim = {d: round(sum(e[d] for e in evs) / len(evs), 2) for d in DIMS}
        summary[s] = {
            "n_evals": len(evs),
            "avg_total": round(avg_total, 2),
            "pct_b1": round((avg_total / 30) * 10, 2),  # B.1 = Cross-eval = 10%
            "by_dim": by_dim,
        }
    return {
        "session": "2026-03-10 — Week 2: First Draft ICP Presentations",
        "evaluations": evals,
        "per_presenter": per_presenter,
        "summary": summary,
        "dimensions": DIMS,
    }


def load_all_classes():
    """Carga component_a.json de cada clase ordenadas por fecha."""
    classes = []
    for sd in sorted(DATA_DIR.iterdir()):
        if not sd.is_dir():
            continue
        ca_file = sd / "component_a.json"
        if not ca_file.exists():
            continue
        try:
            with open(ca_file, encoding="utf-8") as f:
                data = json.load(f)
            data["session_id"] = sd.name
            classes.append(data)
        except Exception as e:
            print(f"⚠️  No pude leer {ca_file}: {e}")
    return classes


def aggregate(classes):
    """Calcula acumulados por alumno."""
    per_student = {s: {
        "n_classes": 0,
        "preparation_sum": 0.0,
        "participation_a_sum": 0.0,
        "participation_b_sum": 0.0,
        "attendance_sum": 0.0,
        "total_a_sum": 0.0,
        "total_b_sum": 0.0,
        "by_class": [],
    } for s in STUDENTS}

    for c in classes:
        sid = c["session_id"]
        summary = c.get("summary", {})
        for s in STUDENTS:
            sm = summary.get(s)
            if not sm:
                continue
            per_student[s]["n_classes"] += 1
            per_student[s]["preparation_sum"] += sm["preparation_pct"]
            per_student[s]["participation_a_sum"] += sm["participation_pct_A_quality_only"]
            per_student[s]["participation_b_sum"] += sm["participation_pct_B_with_bonus"]
            per_student[s]["attendance_sum"] += sm["attendance_pct"]
            per_student[s]["total_a_sum"] += sm["component_a_total_scenario_A"]
            per_student[s]["total_b_sum"] += sm["component_a_total_scenario_B"]
            per_student[s]["by_class"].append({
                "session_id": sid,
                "preparation": sm["preparation_pct"],
                "participation_a": sm["participation_pct_A_quality_only"],
                "participation_b": sm["participation_pct_B_with_bonus"],
                "attendance": sm["attendance_pct"],
                "total_a": sm["component_a_total_scenario_A"],
                "total_b": sm["component_a_total_scenario_B"],
            })

    # Promedios
    for s, d in per_student.items():
        n = d["n_classes"] or 1
        d["preparation_avg"] = round(d["preparation_sum"] / n, 2)
        d["participation_a_avg"] = round(d["participation_a_sum"] / n, 2)
        d["participation_b_avg"] = round(d["participation_b_sum"] / n, 2)
        d["attendance_avg"] = round(d["attendance_sum"] / n, 2)
        d["total_a_avg"] = round(d["total_a_sum"] / n, 2)
        d["total_b_avg"] = round(d["total_b_sum"] / n, 2)
        d["pct_a_avg"] = round(d["total_a_avg"] / 20 * 100, 1)
        d["pct_b_avg"] = round(d["total_b_avg"] / 20 * 100, 1)

    return per_student


def write_json(per_student, classes):
    """JSON consumible por el dashboard."""
    out = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "students": STUDENTS,
        "n_classes": len(classes),
        "class_ids": [c["session_id"] for c in classes],
        "aggregated": per_student,
        "rubric_levels": {
            "6": "Comes prepared. Contributes readily without dominating. Thoughtful, advances conversation. Respects others. Active in small groups.",
            "5": "Comes prepared, makes thoughtful comments when called. Contributes occasionally without prompting. Respects others. Active in small groups.",
            "4": "Generally prepared. Participates but may dominate, ramble, interrupt with digressive questions, bluff when unprepared, miss social cues.",
            "3": "Comes prepared. Does NOT voluntarily contribute. Minimal answers when called. Listens attentively, takes notes.",
            "2": "Comes but NOT prepared. Does not contribute. Unable to contribute usefully when called.",
            "1": "Comes but NOT prepared. May be disruptive. Negative impact on others.",
        },
    }
    out_file = OUT_DIR / "all_classes.json"
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"✅ JSON acumulado: {out_file.relative_to(REPO_ROOT)}")
    return out


def write_excel(per_student, classes):
    """Excel acumulado con hojas: Resumen, Por clase, Detalle por alumno."""
    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="305496")
    SUBHEADER_FILL = PatternFill("solid", fgColor="8EA9DB")
    A_FILL = PatternFill("solid", fgColor="FFF2CC")
    B_FILL = PatternFill("solid", fgColor="C6EFCE")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WHITE = Font(color="FFFFFF", bold=True)
    BOLD = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Aggregated summary
    ws = wb.active
    ws.title = "Aggregated"
    ws.cell(row=1, column=1, value="COMPONENT A — Aggregated across all classes").font = Font(bold=True, size=16, color="305496")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} · {len(classes)} classes").font = Font(italic=True, color="595959")

    headers = ["Student", "Classes", "Prep avg", "Att avg",
               "─ Scen A ─", "Part A avg", "TOTAL A avg", "% A avg",
               "─ Scen B ─", "Part B avg", "TOTAL B avg", "% B avg"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = WHITE; c.fill = SUBHEADER_FILL; c.alignment = CENTER; c.border = BORDER

    sorted_students = sorted(per_student.items(), key=lambda x: x[1]["total_a_avg"], reverse=True)
    row = 5
    for s, d in sorted_students:
        vals = [s, d["n_classes"], d["preparation_avg"], d["attendance_avg"],
                "", d["participation_a_avg"], d["total_a_avg"], d["pct_a_avg"],
                "", d["participation_b_avg"], d["total_b_avg"], d["pct_b_avg"]]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.alignment = CENTER; c.border = BORDER
            if col == 1: c.font = BOLD
            if col == 7: c.fill = A_FILL; c.font = BOLD
            if col == 11: c.fill = B_FILL; c.font = BOLD
            if col in (3, 4, 6, 7, 10, 11): c.number_format = "0.00"
            if col in (8, 12): c.number_format = "0.0\"%\""
        row += 1

    ws.column_dimensions["A"].width = 12
    for i in range(2, 13):
        ws.column_dimensions[get_column_letter(i)].width = 12

    # Sheet 2: Per class — student × class matrix (Scenario A totals)
    ws2 = wb.create_sheet("Per class (Scen A)")
    ws2.cell(row=1, column=1, value="Component A Total per student and class — Scenario A").font = Font(bold=True, size=14, color="305496")
    ws2.cell(row=3, column=1, value="Student").font = WHITE
    ws2.cell(row=3, column=1).fill = SUBHEADER_FILL
    for j, c in enumerate(classes):
        cell = ws2.cell(row=3, column=2 + j, value=c["session_id"])
        cell.font = WHITE; cell.fill = SUBHEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    ws2.cell(row=3, column=2 + len(classes), value="Average").font = WHITE
    ws2.cell(row=3, column=2 + len(classes)).fill = SUBHEADER_FILL
    ws2.cell(row=3, column=2 + len(classes)).border = BORDER

    for i, (s, d) in enumerate(sorted_students):
        ws2.cell(row=4 + i, column=1, value=s).font = BOLD
        ws2.cell(row=4 + i, column=1).border = BORDER
        for j, c in enumerate(classes):
            sm = c.get("summary", {}).get(s)
            v = sm["component_a_total_scenario_A"] if sm else None
            cell = ws2.cell(row=4 + i, column=2 + j, value=v)
            cell.alignment = CENTER; cell.border = BORDER
            cell.number_format = "0.00"
        avg_cell = ws2.cell(row=4 + i, column=2 + len(classes), value=d["total_a_avg"])
        avg_cell.font = BOLD; avg_cell.fill = A_FILL; avg_cell.alignment = CENTER
        avg_cell.border = BORDER; avg_cell.number_format = "0.00"

    ws2.column_dimensions["A"].width = 12
    for i in range(2, 2 + len(classes) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 13

    # Sheet 3: Per class — Scenario B
    ws3 = wb.create_sheet("Per class (Scen B)")
    ws3.cell(row=1, column=1, value="Component A Total per student and class — Scenario B (with bonus)").font = Font(bold=True, size=14, color="305496")
    ws3.cell(row=3, column=1, value="Student").font = WHITE
    ws3.cell(row=3, column=1).fill = SUBHEADER_FILL
    for j, c in enumerate(classes):
        cell = ws3.cell(row=3, column=2 + j, value=c["session_id"])
        cell.font = WHITE; cell.fill = SUBHEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    ws3.cell(row=3, column=2 + len(classes), value="Average").font = WHITE
    ws3.cell(row=3, column=2 + len(classes)).fill = SUBHEADER_FILL

    for i, (s, d) in enumerate(sorted_students):
        ws3.cell(row=4 + i, column=1, value=s).font = BOLD
        ws3.cell(row=4 + i, column=1).border = BORDER
        for j, c in enumerate(classes):
            sm = c.get("summary", {}).get(s)
            v = sm["component_a_total_scenario_B"] if sm else None
            cell = ws3.cell(row=4 + i, column=2 + j, value=v)
            cell.alignment = CENTER; cell.border = BORDER
            cell.number_format = "0.00"
        avg_cell = ws3.cell(row=4 + i, column=2 + len(classes), value=d["total_b_avg"])
        avg_cell.font = BOLD; avg_cell.fill = B_FILL; avg_cell.alignment = CENTER
        avg_cell.border = BORDER; avg_cell.number_format = "0.00"

    ws3.column_dimensions["A"].width = 12
    for i in range(2, 2 + len(classes) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 13

    out_file = OUT_DIR / "Component_A_Aggregated.xlsx"
    wb.save(out_file)
    print(f"✅ Excel acumulado: {out_file.relative_to(REPO_ROOT)}")
    return out_file


def write_dashboard_html(per_student, classes, agg_data):
    """Dashboard standalone HTML que carga datos embebidos."""
    sorted_students = sorted(per_student.items(), key=lambda x: x[1]["total_a_avg"], reverse=True)
    embedded = json.dumps(agg_data, ensure_ascii=False)
    classes_summary = []
    for c in classes:
        cs = {
            "session_id": c["session_id"],
            "n_questions": len(c.get("preparation", {}).get("rows", [])),
            "n_contributions": len(c.get("participation", {}).get("rows", [])),
            "summary": c.get("summary", {}),
        }
        classes_summary.append(cs)
    embedded_classes = json.dumps(classes_summary, ensure_ascii=False)
    embedded_b = json.dumps(get_component_b_data(), ensure_ascii=False)
    embedded_c = json.dumps(get_component_c_data(), ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Global Technology Commercialization (GDI.60030) — Professor Ileana Palaco</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: #f1f5f9; color: #0f172a; line-height: 1.5;
  }}
  .header {{
    background: linear-gradient(135deg, #305496 0%, #1e3a8a 100%);
    color: white; padding: 28px 40px;
  }}
  .header h1 {{ font-size: 26px; margin-bottom: 4px; }}
  .header .subtitle {{ opacity: 0.95; font-size: 15px; font-weight: 500; margin-bottom: 8px; }}
  .header p {{ opacity: 0.8; font-size: 13px; }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 32px 40px; }}
  .toggle-row {{
    display: flex; align-items: center; gap: 16px;
    background: white; padding: 16px 24px; border-radius: 12px;
    margin-bottom: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.05);
  }}
  .toggle-row label {{ font-weight: 600; }}
  .toggle-row select {{
    padding: 8px 12px; border-radius: 8px; border: 1px solid #cbd5e1;
    font-size: 14px; background: white; cursor: pointer;
  }}
  .stats-grid {{
    display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px; margin-bottom: 32px;
  }}
  .stat-card {{
    background: white; padding: 20px 24px; border-radius: 12px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
  }}
  .stat-card .label {{ color: #64748b; font-size: 12px; text-transform: uppercase; letter-spacing: 0.04em; }}
  .stat-card .value {{ font-size: 32px; font-weight: 700; color: #305496; margin-top: 4px; }}
  .stat-card .sublabel {{ color: #94a3b8; font-size: 12px; margin-top: 4px; }}
  .section {{
    background: white; border-radius: 14px; padding: 28px;
    margin-bottom: 28px; box-shadow: 0 1px 3px rgba(0,0,0,0.05);
  }}
  .section h2 {{ font-size: 18px; margin-bottom: 20px; color: #1e293b; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  th {{
    text-align: left; padding: 12px 14px; background: #f8fafc;
    font-weight: 600; color: #475569; border-bottom: 1px solid #e2e8f0;
    font-size: 12px; text-transform: uppercase; letter-spacing: 0.04em;
  }}
  td {{ padding: 14px; border-bottom: 1px solid #f1f5f9; }}
  tr:hover {{ background: #f8fafc; }}
  .student-name {{ font-weight: 600; color: #0f172a; }}
  .pct-bar {{
    background: #e2e8f0; height: 8px; border-radius: 4px; overflow: hidden;
    width: 120px; display: inline-block; vertical-align: middle; margin-right: 8px;
  }}
  .pct-bar > div {{ background: #305496; height: 100%; }}
  .total {{ font-weight: 700; color: #305496; }}
  .badge {{
    display: inline-block; padding: 2px 8px; border-radius: 12px;
    font-size: 11px; font-weight: 600;
  }}
  .badge-rank-1 {{ background: #fef3c7; color: #92400e; }}
  .badge-rank-2 {{ background: #dbeafe; color: #1e40af; }}
  .badge-rank-3 {{ background: #ede9fe; color: #6b21a8; }}
  .badge-rank-other {{ background: #f1f5f9; color: #475569; }}
  .classes-grid {{
    display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 16px;
  }}
  .class-card-link {{ text-decoration: none; color: inherit; }}
  .class-card {{
    background: white; border: 1px solid #e2e8f0; border-radius: 12px;
    padding: 20px; cursor: pointer; transition: all 0.2s; height: 100%;
  }}
  .class-card-link:hover .class-card {{
    border-color: #305496; box-shadow: 0 4px 12px rgba(48, 84, 150, 0.1);
    transform: translateY(-2px);
  }}
  .card-arrow {{ font-size: 12px; color: #305496; font-weight: 600; margin-top: 8px; }}
  .rubric-table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
  .rubric-table th {{ background: #305496; color: white; padding: 12px; text-align: left; font-size: 12px; }}
  .rubric-table td {{ padding: 14px; border-bottom: 1px solid #e2e8f0; vertical-align: top; line-height: 1.7; font-size: 13px; }}
  .rubric-table .score-cell {{ font-size: 28px; font-weight: 700; text-align: center; width: 60px; }}
  .rubric-table tr.rubric-6 {{ background: #d1fae5; }}
  .rubric-table tr.rubric-6 .score-cell {{ color: #065f46; }}
  .rubric-table tr.rubric-5 {{ background: #ecfdf5; }}
  .rubric-table tr.rubric-5 .score-cell {{ color: #047857; }}
  .rubric-table tr.rubric-4 {{ background: #fef3c7; }}
  .rubric-table tr.rubric-4 .score-cell {{ color: #92400e; }}
  .rubric-table tr.rubric-3 {{ background: #fed7aa; }}
  .rubric-table tr.rubric-3 .score-cell {{ color: #9a3412; }}
  .rubric-table tr.rubric-2 {{ background: #fee2e2; }}
  .rubric-table tr.rubric-2 .score-cell {{ color: #991b1b; }}
  .rubric-table tr.rubric-1 {{ background: #fecaca; }}
  .rubric-table tr.rubric-1 .score-cell {{ color: #7f1d1d; }}
  .rubric-rules {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  .rubric-rules th {{ background: #f1f5f9; padding: 10px 14px; text-align: left; font-weight: 600; color: #475569; }}
  .rubric-rules td {{ padding: 10px 14px; border-bottom: 1px solid #f1f5f9; }}
  .calendar-table, .attendance-table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px; }}
  .calendar-table th, .attendance-table th {{ background: #305496; color: white; padding: 10px 14px; text-align: left; font-size: 12px; }}
  .calendar-table td, .attendance-table td {{ padding: 10px 14px; border-bottom: 1px solid #f1f5f9; }}
  .calendar-table tr.recorded {{ background: #ecfdf5; }}
  .calendar-table tr.virtual {{ background: #fef3c7; }}
  .calendar-table tr.noclass {{ background: #fee2e2; }}
  .calendar-table a {{ color: #305496; font-weight: 600; text-decoration: none; }}
  .calendar-table a:hover {{ text-decoration: underline; }}
  .status-pill {{ display: inline-block; padding: 3px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; }}
  .status-rec {{ background: #d1fae5; color: #065f46; }}
  .status-notrec {{ background: #f1f5f9; color: #475569; }}
  .status-virt {{ background: #fef3c7; color: #92400e; }}
  .status-noclass {{ background: #fecaca; color: #7f1d1d; }}
  .status-pending {{ background: #ede9fe; color: #5b21b6; }}
  .attendance-table td.ok {{ color: #065f46; font-weight: 700; text-align: center; background: #d1fae5; }}
  .attendance-table td.absent {{ color: #7f1d1d; font-weight: 700; text-align: center; background: #fecaca; }}
  .attendance-table td.absent-permission {{ color: #92400e; font-weight: 700; text-align: center; background: #fef3c7; font-size: 11px; }}
  .attendance-table td.pending {{ color: #cbd5e1; font-weight: 700; text-align: center; background: #fafafa; }}
  .attendance-table td.total {{ font-weight: 700; text-align: center; background: #f8fafc; }}
  .attendance-table tr.pending-row td:first-child {{ color: #94a3b8; font-style: italic; }}
  .collapsible-header {{
    cursor: pointer; user-select: none; display: flex; align-items: center;
    justify-content: space-between; transition: color 0.2s;
  }}
  .collapsible-header:hover {{ color: #305496; }}
  .toggle-icon {{
    display: inline-block; font-size: 14px; color: #305496; transition: transform 0.2s;
    width: 24px; height: 24px; line-height: 24px; text-align: center;
    background: #f1f5f9; border-radius: 50%;
  }}
  .toggle-icon.open {{ transform: rotate(90deg); }}
  .collapsible-header:hover .toggle-icon {{ background: #dbeafe; }}
  .policy-table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px; }}
  .policy-table th {{
    background: #305496; color: white; padding: 10px 12px; text-align: left;
    font-size: 12px; text-transform: uppercase; letter-spacing: 0.04em;
  }}
  .policy-table td {{ padding: 10px 12px; border-bottom: 1px solid #e2e8f0; vertical-align: middle; }}
  .policy-letter {{
    font-size: 22px; font-weight: 800; text-align: center;
    background: #f1f5f9; color: #475569;
  }}
  .policy-desc {{ font-weight: 600; color: #1e293b; }}
  .policy-weight {{ font-weight: 700; text-align: center; color: #305496; background: #f8fafc; }}
  .policy-pct {{ font-weight: 600; text-align: center; color: #64748b; }}
  .policy-A-main {{ background: #dbeafe !important; }}
  .policy-A-main .policy-letter {{ background: #305496 !important; color: white; font-size: 26px; }}
  .policy-A-main .policy-desc {{ color: #1e3a8a; font-weight: 700; }}
  .policy-A-main .policy-weight {{ background: #93c5fd; color: #1e3a8a; }}
  .policy-A-main td {{ border-bottom-color: #93c5fd; }}
  .policy-total {{ background: #1e293b; color: white; }}
  .policy-total td {{ color: white !important; border: none; background: #1e293b !important; }}
  .policy-total .policy-weight {{ color: white !important; background: #1e293b !important; }}
  .main-tabs {{
    display: grid; grid-template-columns: repeat(5, 1fr); gap: 8px;
    margin-bottom: 24px;
  }}
  .main-tab {{
    background: white; border: 2px solid #e2e8f0; border-radius: 12px;
    padding: 14px 12px; cursor: pointer; text-align: left;
    display: flex; align-items: center; gap: 12px;
    transition: all 0.2s; font-family: inherit;
  }}
  .main-tab:hover {{
    border-color: #305496; transform: translateY(-2px);
    box-shadow: 0 4px 12px rgba(48, 84, 150, 0.1);
  }}
  .main-tab.active {{
    background: linear-gradient(135deg, #305496 0%, #1e3a8a 100%);
    border-color: #305496; color: white;
  }}
  .main-tab.active .tab-letter {{ background: white; color: #305496; }}
  .main-tab.active .tab-text small {{ color: #c7d7f0; }}
  .main-tab.disabled {{ opacity: 0.6; }}
  .main-tab.disabled:hover {{ transform: none; box-shadow: none; }}
  .tab-letter {{
    width: 36px; height: 36px; border-radius: 50%;
    background: #305496; color: white;
    display: flex; align-items: center; justify-content: center;
    font-weight: 800; font-size: 18px; flex-shrink: 0;
  }}
  .tab-text {{ font-size: 12px; line-height: 1.4; font-weight: 600; }}
  .tab-text small {{ font-weight: 500; color: #94a3b8; font-size: 11px; }}
  .main-tab-content {{ display: none; }}
  .main-tab-content.active {{ display: block; }}
  .coming-soon {{ text-align: center; padding: 48px 24px; }}
  .coming-soon h2 {{ color: #305496; margin-bottom: 12px; }}
  .coming-soon p {{ color: #64748b; max-width: 500px; margin: 0 auto 24px; }}
  .cs-badge {{
    display: inline-block; background: #fef3c7; color: #92400e;
    padding: 10px 24px; border-radius: 30px; font-weight: 700; font-size: 14px;
  }}
  .b-matrix {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  .b-matrix th {{ background: #305496; color: white; padding: 10px 12px; text-align: center; }}
  .b-matrix th:first-child {{ text-align: left; }}
  .b-matrix td {{ padding: 10px 12px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600; }}
  .b-matrix td.cell-high {{ background: #d1fae5; color: #065f46; }}
  .b-matrix td.cell-mid {{ background: #fef3c7; color: #92400e; }}
  .b-matrix td.cell-low {{ background: #fee2e2; color: #991b1b; }}
  .b-matrix td.cell-self {{ background: #f1f5f9; color: #94a3b8; }}
  .b-matrix td.cell-empty {{ background: #fafafa; color: #cbd5e1; }}
  .b-summary-table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px; }}
  .b-summary-table th {{ background: #305496; color: white; padding: 12px 14px; text-align: left; }}
  .b-summary-table th small {{ font-weight: 500; opacity: 0.85; }}
  .b-summary-table td {{ padding: 14px; border-bottom: 1px solid #e2e8f0; }}
  .b-summary-table tr:hover {{ background: #f8fafc; }}
  .b-cell-done {{ background: #ecfdf5; color: #065f46; font-weight: 600; text-align: center; }}
  .b-cell-pending {{ background: #fef3c7; color: #92400e; font-style: italic; text-align: center; }}
  .b-pending-table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px; }}
  .b-pending-table th {{ background: #305496; color: white; padding: 10px 12px; text-align: center; }}
  .b-pending-table th:first-child {{ text-align: left; }}
  .b-pending-table td {{ padding: 10px 12px; border: 1px solid #e2e8f0; text-align: center; }}
  .b-pending-table td.cell-empty {{ background: #fafafa; color: #cbd5e1; font-weight: 600; }}
  .form-buttons {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 8px; }}
  .form-btn {{
    display: flex; align-items: center; gap: 14px;
    padding: 16px 20px; border-radius: 12px; text-decoration: none;
    border: 2px solid; transition: all 0.2s;
  }}
  .form-btn:hover {{ transform: translateY(-2px); box-shadow: 0 6px 14px rgba(0,0,0,0.08); }}
  .form-btn-b1 {{ background: #ecfdf5; border-color: #10b981; color: #065f46; }}
  .form-btn-b1:hover {{ background: #d1fae5; border-color: #047857; }}
  .form-btn-b2 {{ background: #eff6ff; border-color: #305496; color: #1e3a8a; }}
  .form-btn-b2:hover {{ background: #dbeafe; border-color: #1e3a8a; }}
  .form-btn-icon {{ font-size: 28px; }}
  .form-btn-text {{ flex: 1; display: flex; flex-direction: column; gap: 2px; }}
  .form-btn-text strong {{ font-size: 14px; }}
  .form-btn-text small {{ font-size: 12px; opacity: 0.85; }}
  .form-btn-arrow {{ font-size: 22px; font-weight: 700; opacity: 0.5; }}
  .form-btn:hover .form-btn-arrow {{ opacity: 1; }}
  @media (max-width: 720px) {{ .form-buttons {{ grid-template-columns: 1fr; }} }}
  .exam-card {{
    background: white; border: 1px solid #e2e8f0; border-radius: 12px;
    margin-bottom: 16px; overflow: hidden;
  }}
  .exam-header {{
    padding: 14px 20px; margin: 0; font-size: 15px; font-weight: 600;
    background: #f8fafc; border-bottom: 1px solid #e2e8f0;
  }}
  .exam-header:hover {{ background: #f1f5f9; }}
  .exam-content {{ padding: 16px 20px; }}
  .student-block {{
    margin-bottom: 14px; border: 1px solid #e2e8f0; border-radius: 8px;
    overflow: hidden;
  }}
  .student-header {{
    padding: 12px 16px; margin: 0; background: #eff6ff;
    border-bottom: 1px solid #dbeafe; font-size: 14px; font-weight: 600;
  }}
  .student-header:hover {{ background: #dbeafe; }}
  .student-answers {{ padding: 16px 20px; }}
  .answer-block {{
    padding: 16px; margin-bottom: 14px;
    background: #f8fafc; border-left: 4px solid #305496; border-radius: 6px;
  }}
  .answer-block:last-child {{ margin-bottom: 0; }}
  .question-text {{ font-size: 16px; color: #1e293b; margin-bottom: 10px; }}
  .question-text strong {{ font-weight: 700; }}
  .answer-text {{
    font-size: 14px; color: #334155; line-height: 1.65;
    background: white; padding: 12px; border-radius: 6px;
    margin-bottom: 10px; white-space: pre-wrap;
  }}
  .answer-score {{
    display: flex; align-items: center; gap: 10px;
    padding: 8px 12px; background: white; border-radius: 6px; border: 1px solid #e2e8f0;
  }}
  .answer-score label {{ font-weight: 600; color: #475569; font-size: 13px; }}
  .score-input {{
    width: 70px; padding: 6px 8px; font-size: 14px; font-weight: 700;
    border: 2px solid #305496; border-radius: 6px; text-align: center;
    color: #1e3a8a; background: #f0f9ff;
  }}
  .score-input:disabled {{ background: #ecfdf5; border-color: #10b981; color: #065f46; cursor: not-allowed; }}
  .score-max {{ font-weight: 600; color: #64748b; }}
  .class-card .date {{ font-size: 16px; font-weight: 700; color: #305496; }}
  .class-card .meta {{ font-size: 13px; color: #64748b; margin-top: 4px; }}
  .class-card .stats-mini {{ display: flex; gap: 12px; margin-top: 12px; font-size: 12px; }}
  .class-card .stats-mini span {{ color: #475569; }}
  .chart-wrapper {{ position: relative; height: 350px; }}
  .footer {{ text-align: center; padding: 32px; color: #94a3b8; font-size: 12px; }}
</style>
</head>
<body>

<div class="header">
  <h1>📊 Global Technology Commercialization (GDI.60030)</h1>
  <p class="subtitle">Professor Ileana Palaco · TA Cesar Fonseca</p>
  <p>Component A · Preparation 7.5% + Participation 7.5% + Attendance 5% = 20% of total grade</p>
</div>

<div class="container">

  <select id="scenarioSelect" style="display:none">
    <option value="A" selected>A</option>
  </select>

  <div class="section">
    <h2 class="collapsible-header" onclick="toggleSection('gradePolicyFull', 'gradePolicyToggle')">
      <span>📑 Grade Policy — Full Course Breakdown</span>
      <span class="toggle-icon" id="gradePolicyToggle">▶</span>
    </h2>
    <p style="color:#64748b; font-size:13px; margin-bottom:16px;">
      Complete grading structure for the course. <strong>This dashboard covers Component A only</strong> (Preparation, Participation & Attendance — 20% of total grade).
    </p>
    <div id="gradePolicyFull" style="display:none;">
      <table class="policy-table">
        <thead>
          <tr>
            <th style="width:40px">Component</th>
            <th>Description</th>
            <th style="width:80px">Weight</th>
            <th>Detailed Composition</th>
            <th style="width:60px">%</th>
            <th>Notes</th>
          </tr>
        </thead>
        <tbody>
          <tr class="policy-A-main">
            <td rowspan="3" class="policy-letter">A</td>
            <td rowspan="3" class="policy-desc">Student's Preparation, Participation &amp; Attendance</td>
            <td rowspan="3" class="policy-weight">20%</td>
            <td><strong>A1</strong> Student's Preparation</td><td class="policy-pct">7.5%</td><td>Questions to students</td>
          </tr>
          <tr class="policy-A-main"><td><strong>A2</strong> Student's Participation</td><td class="policy-pct">7.5%</td><td>Questions to Professor</td></tr>
          <tr class="policy-A-main"><td><strong>A3</strong> Student's Attendance</td><td class="policy-pct">5.0%</td><td>—</td></tr>

          <tr>
            <td rowspan="3" class="policy-letter">B</td>
            <td rowspan="3" class="policy-desc">Individual Classwork / Worksheet / Team discussion</td>
            <td rowspan="3" class="policy-weight">25%</td>
            <td><strong>B1</strong> Individual Classwork</td><td class="policy-pct">10.0%</td><td>—</td>
          </tr>
          <tr><td><strong>B2</strong> Worksheet</td><td class="policy-pct">10.0%</td><td>—</td></tr>
          <tr><td><strong>B3</strong> Team discussion</td><td class="policy-pct">5.0%</td><td>—</td></tr>

          <tr>
            <td rowspan="2" class="policy-letter">C</td>
            <td rowspan="2" class="policy-desc">Assignments &amp; Partial Submissions</td>
            <td rowspan="2" class="policy-weight">20%</td>
            <td><strong>C1</strong> Assignments</td><td class="policy-pct">10.0%</td><td>1) Popup Quiz Reflection, 2) GTC Framework process</td>
          </tr>
          <tr><td><strong>C2</strong> Partial Submissions</td><td class="policy-pct">10.0%</td><td>ICP (V1-V4)</td></tr>

          <tr>
            <td class="policy-letter">D</td>
            <td class="policy-desc">Final Term Submission Report</td>
            <td class="policy-weight">15%</td>
            <td><strong>D1</strong></td><td class="policy-pct">15.0%</td><td>Based on ICP Evaluation Rubric</td>
          </tr>

          <tr>
            <td class="policy-letter">E</td>
            <td class="policy-desc">Final GTC Project Pitch</td>
            <td class="policy-weight">20%</td>
            <td><strong>E1</strong></td><td class="policy-pct">20.0%</td><td>Based on all the feedback provided by professor during the course. All professor guidelines must be followed (references, PPT slides, formatting, etc.)</td>
          </tr>

          <tr class="policy-total">
            <td colspan="2" style="text-align:right; font-weight:700;">TOTAL</td>
            <td class="policy-weight">100%</td>
            <td colspan="3"></td>
          </tr>
        </tbody>
      </table>
      <p style="margin-top:12px; padding:12px; background:#dbeafe; border-left:4px solid #305496; font-size:13px; color:#1e3a8a;">
        🔵 <strong>Component A</strong> (highlighted in blue above) is what this dashboard tracks and grades automatically based on class transcripts and the 1–6 rubric.
      </p>
    </div>
  </div>

  <!-- Top-level tabs by Component (A-E) -->
  <div class="main-tabs">
    <button class="main-tab active" onclick="showMainTab(this, 'A')">
      <span class="tab-letter">A</span>
      <span class="tab-text">Preparation, Participation &amp; Attendance<br><small>20% · Active</small></span>
    </button>
    <button class="main-tab disabled" onclick="showMainTab(this, 'B')">
      <span class="tab-letter">B</span>
      <span class="tab-text">Individual Classwork / Worksheet / Team discussion<br><small>25% · Coming soon</small></span>
    </button>
    <button class="main-tab disabled" onclick="showMainTab(this, 'C')">
      <span class="tab-letter">C</span>
      <span class="tab-text">Assignments &amp; Partial Submissions<br><small>20% · Coming soon</small></span>
    </button>
    <button class="main-tab disabled" onclick="showMainTab(this, 'D')">
      <span class="tab-letter">D</span>
      <span class="tab-text">Final Term Submission Report<br><small>15% · Coming soon</small></span>
    </button>
    <button class="main-tab disabled" onclick="showMainTab(this, 'E')">
      <span class="tab-letter">E</span>
      <span class="tab-text">Final GTC Project Pitch<br><small>20% · Coming soon</small></span>
    </button>
  </div>

  <!-- TAB A CONTENT (active) -->
  <div id="main-tab-A" class="main-tab-content active">

    <div class="section">
      <h2 style="color:#1e293b;">📊 Component A — Preparation, Participation &amp; Attendance (20%)</h2>
      <p style="color:#64748b; font-size:13px;">Preparation 7.5% (Questions to students) + Participation 7.5% (Questions to Professor) + Attendance 5%</p>
    </div>

  <div class="stats-grid" id="statsGrid"></div>

  <div class="section">
    <h2>🏆 Aggregated student ranking</h2>
    <table id="rankingTable"></table>
  </div>

  <div class="section">
    <h2>📈 Total evolution by class</h2>
    <div class="chart-wrapper"><canvas id="evolutionChart"></canvas></div>
  </div>

  <div class="section">
    <h2 class="collapsible-header" onclick="toggleSection('classDetailContent', 'classDetailToggle')">
      <span>📅 Class detail</span>
      <span class="toggle-icon open" id="classDetailToggle">▶</span>
    </h2>
    <div id="classDetailContent" style="display:block;">
      <div class="classes-grid" id="classesGrid"></div>
    </div>
  </div>

  <div class="section">
    <h2 class="collapsible-header" onclick="toggleSection('calendarFull', 'calendarToggle')">
      <span>📅 Class Attendance Calendar — Semester Overview</span>
      <span class="toggle-icon" id="calendarToggle">▶</span>
    </h2>
    <p style="color:#64748b; font-size:13px; margin-bottom:16px;">
      Weekly summary of recorded sessions, holidays and presentations days. Click the title to expand or collapse.
    </p>
    <div id="calendarFull" style="display:none;">
      <table class="calendar-table">
        <thead><tr><th>Week</th><th>Dates</th><th>Type / Class</th><th>Status</th><th>Detail</th></tr></thead>
        <tbody>
          <tr><td>1</td><td>Tue, Mar 3</td><td>Lecture 1</td><td><span class="status-pill status-notrec">Not recorded</span></td><td>—</td></tr>
          <tr><td>2</td><td>Tue, Mar 10</td><td>Lecture 2 + ICP First Draft Presentations</td><td><span class="status-pill status-notrec">Not recorded</span></td><td>—</td></tr>
          <tr class="recorded"><td>3</td><td>Tue, Mar 17</td><td>Paper Revision</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-03-17.html">View detail →</a></td></tr>
          <tr><td>4</td><td>Tue, Mar 24</td><td>Lecture L4</td><td><span class="status-pill status-notrec">Not recorded</span></td><td>—</td></tr>
          <tr class="recorded"><td>5</td><td>Tue, Mar 31</td><td>Class session</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-03-31.html">View detail →</a></td></tr>
          <tr class="recorded"><td>6</td><td>Tue, Apr 7</td><td>Class session</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-04-07.html">View detail →</a></td></tr>
          <tr class="recorded"><td>7</td><td>Tue, Apr 14</td><td>GTC Lecture 6 + Presentations</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-04-14.html">View detail →</a></td></tr>
          <tr class="recorded"><td>8</td><td>Tue, Apr 21</td><td>Class session</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-04-21.html">View detail →</a></td></tr>
          <tr><td>9</td><td>Tue, Apr 28</td><td>Lecture L9</td><td><span class="status-pill status-notrec">Not recorded</span></td><td>—</td></tr>
          <tr class="noclass"><td>10</td><td>Tue, May 5</td><td>Holiday (no class) · ICP V4 submission</td><td><span class="status-pill status-noclass">🚫 No class · Holiday</span></td><td>—</td></tr>
          <tr class="recorded"><td>11</td><td>Tue, May 12</td><td>Class session</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-05-12.html">View detail →</a></td></tr>
          <tr class="recorded"><td>12</td><td>Tue, May 19</td><td>Class session</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-05-19.html">View detail →</a></td></tr>
          <tr class="recorded"><td>13</td><td>Tue, May 26</td><td>Class session</td><td><span class="status-pill status-rec">✅ Recorded</span></td><td><a href="classes/2026-05-26.html">View detail →</a></td></tr>
          <tr><td>14</td><td>Tue, Jun 2</td><td><em>Upcoming</em></td><td><span class="status-pill status-pending">⏳ Pending</span></td><td>—</td></tr>
          <tr><td>15</td><td>Tue, Jun 9</td><td><em>Upcoming</em></td><td><span class="status-pill status-pending">⏳ Pending</span></td><td>—</td></tr>
          <tr><td>16</td><td>Tue, Jun 16</td><td><em>Upcoming</em></td><td><span class="status-pill status-pending">⏳ Pending</span></td><td>—</td></tr>
        </tbody>
      </table>

      <h3 style="margin-top:24px; font-size:15px; color:#1e293b;">👥 Attendance by student (recorded classes)</h3>
      <table class="attendance-table">
        <thead><tr><th>Class</th><th>Aryang</th><th>Mega</th><th>Chilaka</th><th>Grace</th><th>Sthepen</th></tr></thead>
        <tbody>
          <tr><td>Tue, Mar 17 (Wk 3)</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td></tr>
          <tr><td>Tue, Mar 31 (Wk 5)</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td></tr>
          <tr><td>Tue, Apr 7 (Wk 6)</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="absent-permission" title="Absent with permission — full attendance credit, no Prep/Part data">⚠ Absent (Permission)</td></tr>
          <tr><td>Tue, Apr 14 (Wk 7) — Presentations</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td></tr>
          <tr><td>Tue, Apr 21 (Wk 8)</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td></tr>
          <tr><td>Tue, May 12 (Wk 11)</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td></tr>
          <tr><td>Tue, May 19 (Wk 12)</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td></tr>
          <tr><td>Tue, May 26 (Wk 13)</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td><td class="ok">✓</td></tr>
          <tr class="pending-row"><td>Tue, Jun 2 (Wk 14)</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td></tr>
          <tr class="pending-row"><td>Tue, Jun 9 (Wk 15)</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td></tr>
          <tr class="pending-row"><td>Tue, Jun 16 (Wk 16)</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td><td class="pending">—</td></tr>
        </tbody>
        <tfoot>
          <tr><td>Total present (recorded)</td><td class="total">8/8</td><td class="total">8/8</td><td class="total">8/8</td><td class="total">8/8</td><td class="total">7/8</td></tr>
        </tfoot>
      </table>
      <p style="color:#94a3b8; font-size:12px; margin-top:10px; font-style:italic;">
        — Empty cells (—) in weeks 14–16 will be updated once classes are held and processed.
      </p>
    </div>
  </div>

  <div class="section">
    <h2 class="collapsible-header" onclick="toggleSection('rubricFull', 'rubricToggle')">
      <span>📋 Evaluation Rubric — Preparation & Participation</span>
      <span class="toggle-icon" id="rubricToggle">▶</span>
    </h2>
    <p style="color:#64748b; font-size:13px; margin-bottom:16px;">
      Official 1–6 scale used to grade each student response and contribution. Click the title to expand or collapse.
    </p>
    <div id="rubricFull" style="display:none;">
      <table class="rubric-table">
        <thead>
          <tr><th style="width:60px">Score</th><th>Descriptor</th></tr>
        </thead>
        <tbody>
          <tr class="rubric-6"><td class="score-cell">6</td>
            <td>• Comes to class prepared<br>
                • Contributes readily to the conversation but doesn't dominate<br>
                • Makes thoughtful contributions that advance the conversation<br>
                • Shows interest in and respect for others' views<br>
                • Participates actively in small groups</td></tr>
          <tr class="rubric-5"><td class="score-cell">5</td>
            <td>• Comes to class prepared and makes thoughtful comments when called upon<br>
                • Contributes occasionally without prompting<br>
                • Shows interest in and respect for others' views<br>
                • Participates actively in small groups</td></tr>
          <tr class="rubric-4"><td class="score-cell">4</td>
            <td>• Generally comes to class prepared<br>
                • Participates in discussion, but may talk too much, make rambling or tangential contributions, continually interrupt the instructor with digressive questions<br>
                • Bluffs their way when unprepared, or otherwise dominates discussions<br>
                • Does not acknowledge cues of annoyance from instructor or other students</td></tr>
          <tr class="rubric-3"><td class="score-cell">3</td>
            <td>• Comes to class prepared<br>
                • Does NOT voluntarily contribute to discussions<br>
                • Gives only minimal answers when called<br>
                • Shows interest in the discussion, listens attentively and takes notes</td></tr>
          <tr class="rubric-2"><td class="score-cell">2</td>
            <td>• Comes to class but has NOT prepared<br>
                • Does not voluntarily contribute to discussions<br>
                • Unlikely to be able to contribute usefully even when called to do so</td></tr>
          <tr class="rubric-1"><td class="score-cell">1</td>
            <td>• Comes to class but has NOT prepared<br>
                • May be disruptive<br>
                • May have a negative impact on others in the group</td></tr>
        </tbody>
      </table>

      <h3 style="margin-top:24px; font-size:15px; color:#1e293b;">📐 Application rules</h3>
      <table class="rubric-rules">
        <thead><tr><th>Situation</th><th>Assigned score</th></tr></thead>
        <tbody>
          <tr><td><strong>Directed</strong> question to student + answered</td><td><strong>1–6</strong> based on quality</td></tr>
          <tr><td><strong>Directed</strong> question + did NOT answer</td><td><strong>0</strong> (penalizes lack of preparation)</td></tr>
          <tr><td><strong>Open</strong> question + student answered</td><td><strong>1–6</strong> based on quality</td></tr>
          <tr><td>Open question + did NOT answer</td><td><strong>N/A</strong> (not applicable)</td></tr>
          <tr><td>Question directed to ANOTHER student</td><td><strong>N/A</strong></td></tr>
          <tr><td>Student's voluntary question or comment</td><td><strong>1–6</strong> based on quality</td></tr>
        </tbody>
      </table>

      <h3 style="margin-top:24px; font-size:15px; color:#1e293b;">🧮 Component A calculation formula (20%)</h3>
      <div style="background:#f8fafc; padding:16px; border-radius:8px; font-family:monospace; font-size:13px; line-height:1.8;">
        <strong>Preparation (7.5%)</strong> = (sum of scores / (n_opportunities × 6)) × 7.5<br>
        <strong>Participation (7.5%)</strong> = (sum of scores / (n_contributions × 6)) × 7.5<br>
        <strong>Attendance (5%)</strong> = manual (5.0 if present, 0 if absent)<br>
        <br>
        <strong>Total A (out of 20)</strong> = Prep + Part + Att
      </div>
    </div>
  </div>

  </div><!-- /main-tab-A -->

  <!-- TAB B CONTENT -->
  <div id="main-tab-B" class="main-tab-content">
    <div class="section">
      <h2 style="color:#1e293b;">📊 Component B — Individual Classwork / Worksheet / Team discussion (25%)</h2>
      <p style="color:#64748b; font-size:13px; margin-bottom:16px;">B.1 Individual Classwork (10%) + B.2 Worksheet (10%) + B.3 Team discussion (5%)</p>

      <div class="form-buttons">
        <a href="https://forms.gle/jXEW3uPj94wppenQ6" target="_blank" rel="noopener" class="form-btn form-btn-b1">
          <span class="form-btn-icon">📝</span>
          <span class="form-btn-text">
            <strong>Student's Cross-evaluation Form</strong>
            <small>B.1 · For students to evaluate peers (10%)</small>
          </span>
          <span class="form-btn-arrow">→</span>
        </a>
        <a href="https://forms.gle/CGCHn4iz1KwQrKaV6" target="_blank" rel="noopener" class="form-btn form-btn-b2">
          <span class="form-btn-icon">🎓</span>
          <span class="form-btn-text">
            <strong>Professor &amp; TA's Evaluation Form</strong>
            <small>B.2 · Internal form for Prof. &amp; TA (15%)</small>
          </span>
          <span class="form-btn-arrow">→</span>
        </a>
      </div>
    </div>

    <div class="section">
      <h2>🧮 Component B Summary — Status overview per student</h2>
      <p style="color:#64748b; font-size:13px; margin-bottom:16px;">
        Combined progress on B.1 (already received) and B.2 (pending Professor's evaluation).
      </p>
      <table id="bSummaryTable" class="b-summary-table"></table>
      <p style="margin-top:12px; padding:12px; background:#fef3c7; border-left:4px solid #f59e0b; font-size:13px; color:#78350f;">
        ⏳ <strong>B.2 (15%) is pending</strong> — once Professor Ileana provides her scores, the table below will be updated and totals here will recalculate automatically.
      </p>
    </div>

    <div class="section">
      <h2>🎤 B.1 — Student's Cross-evaluation (10%)</h2>
      <p style="color:#64748b; font-size:13px; margin-bottom:16px;">
        Session: <strong id="b1Session"></strong>
      </p>

      <div class="stats-grid" id="b1Stats"></div>

      <h3 style="margin-top:24px; font-size:15px; color:#1e293b;">🏆 Cross-evaluation Ranking</h3>
      <table id="b1RankingTable" style="margin-top:8px"></table>

      <h3 style="margin-top:24px; font-size:15px; color:#1e293b;">📊 Average score per dimension (Bar chart)</h3>
      <div class="chart-wrapper" style="height: 380px; margin-top: 12px;"><canvas id="b1DimChart"></canvas></div>

      <h3 style="margin-top:24px; font-size:15px; color:#1e293b;">📋 Evaluation Matrix (evaluator × presenter)</h3>
      <p style="color:#64748b; font-size:12px;">Total score (sum of 5 dimensions, max 30 per evaluation). Hover for breakdown.</p>
      <table id="b1Matrix" class="b-matrix"></table>

      <h3 style="margin-top:24px; font-size:15px; color:#1e293b;">💬 Comments received</h3>
      <table id="b1Comments" style="margin-top:8px"></table>
    </div>

    <div class="section">
      <h2>🎓 B.2 — Professor &amp; TA's Evaluation (15%)</h2>
      <p style="color:#64748b; font-size:13px; margin-bottom:8px;">
        Internal Google Form with extra questions. Scores follow the same 1–6 scale across 5 dimensions
        (or custom scoring decided by Prof. Ileana).
      </p>
      <p style="padding:12px; background:#dbeafe; border-left:4px solid #305496; font-size:13px; color:#1e3a8a; margin:12px 0;">
        📝 <strong>How to fill this:</strong> Once Prof. Ileana &amp; TA provide the scores, the cells below
        (marked <code>—</code>) will be replaced with values 1–6 per dimension per presenter.
        Total per student is computed as <code>(sum / 30) × 15</code> for the 15% weight.
      </p>

      <h3 style="margin-top:16px; font-size:15px; color:#1e293b;">Awaiting evaluation matrix</h3>
      <table id="b2PendingTable" class="b-pending-table"></table>

      <p style="color:#94a3b8; font-size:12px; margin-top:10px; font-style:italic;">
        — Empty cells will be filled once scores are provided. The Component B Summary above will reflect the
        final % once B.2 is complete.
      </p>
    </div>
  </div>

  <!-- TAB C CONTENT -->
  <div id="main-tab-C" class="main-tab-content">
    <div class="section">
      <h2 style="color:#1e293b;">📝 Component C — Assignments &amp; Partial Submissions (20%)</h2>
      <p style="color:#64748b; font-size:13px;">C.1 Assignments (10%) + C.2 Partial Submissions (10%)</p>
    </div>

    <div class="section">
      <h2>📚 C.1 Summary — Total per student (so far)</h2>
      <p style="color:#64748b; font-size:13px; margin-bottom:16px;">
        Sum of scores across all 5 exams (max 100 per exam × 5 = 500 total) → % of the 10% weight.
      </p>
      <table id="c1SummaryTable" class="b-summary-table"></table>
    </div>

    <div class="section">
      <h2>📝 C.1 — Assignments Exam-by-exam (10%)</h2>
      <p style="color:#64748b; font-size:13px; margin-bottom:16px;">
        5 in-class assessments throughout the semester. Each exam has 4 questions worth 25 points each (max 100).
        Click an exam to see all student answers and scores.
      </p>
      <div id="c1ExamsList"></div>
    </div>

    <div class="section coming-soon">
      <h3 style="margin:0 0 12px; color:#305496;">📄 C.2 — Partial Submissions (10%)</h3>
      <p>ICP (V1-V4)</p>
      <div class="cs-badge">⏳ Coming soon</div>
    </div>
  </div>

  <!-- TAB D CONTENT -->
  <div id="main-tab-D" class="main-tab-content">
    <div class="section coming-soon">
      <h2>📄 Component D — Final Term Submission Report (15%)</h2>
      <p>Based on ICP Evaluation Rubric.</p>
      <div class="cs-badge">⏳ Coming soon</div>
    </div>
  </div>

  <!-- TAB E CONTENT -->
  <div id="main-tab-E" class="main-tab-content">
    <div class="section coming-soon">
      <h2>🎯 Component E — Final GTC Project Pitch (20%)</h2>
      <p>Based on all the feedback provided by professor during the course. All professor guidelines must be followed (references, PPT slides, formatting, etc.)</p>
      <div class="cs-badge">⏳ Coming soon</div>
    </div>
  </div>

</div>

<div class="footer">
  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} · Official 1–6 rubric (Class Preparation & Participation) · Automatic text mapping.
</div>

<script>
const DATA = {embedded};
const CLASSES = {embedded_classes};
const COMPONENT_B = {embedded_b};
const COMPONENT_C = {embedded_c};
const COLORS = ['#305496', '#16a34a', '#dc2626', '#9333ea', '#ea580c'];
let evolutionChart = null;
let b1DimChart = null;

const nClassesEl = document.getElementById('nClasses');
if (nClassesEl) nClassesEl.textContent = DATA.n_classes;

function toggleSection(contentId, iconId) {{
  const content = document.getElementById(contentId);
  const icon = document.getElementById(iconId);
  const isOpen = content.style.display === 'block';
  content.style.display = isOpen ? 'none' : 'block';
  icon.classList.toggle('open', !isOpen);
}}

function showMainTab(btn, letter) {{
  document.querySelectorAll('.main-tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.main-tab-content').forEach(t => t.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('main-tab-' + letter).classList.add('active');
  // Render del chart de B solo cuando se abre (canvas debe ser visible para Chart.js)
  if (letter === 'B' && !bRendered) {{
    setTimeout(() => {{
      try {{ renderComponentB(false); bRendered = true; }}
      catch(e) {{ console.error('[componentB chart] failed:', e); }}
    }}, 50);
  }}
  // Re-render the chart if switching to A (canvas needs to be in DOM)
  if (letter === 'A' && evolutionChart) {{
    setTimeout(() => evolutionChart.resize(), 50);
  }}
}}

function renderStats() {{
  const scenario = document.getElementById('scenarioSelect').value;
  const totalKey = scenario === 'A' ? 'total_a_avg' : 'total_b_avg';
  const pctKey = scenario === 'A' ? 'pct_a_avg' : 'pct_b_avg';
  const students = Object.entries(DATA.aggregated);
  const groupAvg = students.reduce((s, [_, d]) => s + d[totalKey], 0) / students.length;
  const top = students.sort((a, b) => b[1][totalKey] - a[1][totalKey])[0];
  const bottom = students[students.length - 1];

  document.getElementById('statsGrid').innerHTML = `
    <div class="stat-card">
      <div class="label">Group average</div>
      <div class="value">${{groupAvg.toFixed(2)}}</div>
      <div class="sublabel">/ 20 points · ${{(groupAvg / 20 * 100).toFixed(1)}}%</div>
    </div>
    <div class="stat-card">
      <div class="label">Classes processed</div>
      <div class="value">${{DATA.n_classes}}</div>
      <div class="sublabel">${{DATA.class_ids.join(' · ')}}</div>
    </div>
    <div class="stat-card">
      <div class="label">Top student</div>
      <div class="value">${{top[0]}}</div>
      <div class="sublabel">${{top[1][totalKey].toFixed(2)}} average</div>
    </div>
    <div class="stat-card">
      <div class="label">Needs attention</div>
      <div class="value">${{bottom[0]}}</div>
      <div class="sublabel">${{bottom[1][totalKey].toFixed(2)}} average</div>
    </div>
  `;
}}

function renderRanking() {{
  const scenario = document.getElementById('scenarioSelect').value;
  const totalKey = scenario === 'A' ? 'total_a_avg' : 'total_b_avg';
  const partKey = scenario === 'A' ? 'participation_a_avg' : 'participation_b_avg';
  const pctKey = scenario === 'A' ? 'pct_a_avg' : 'pct_b_avg';

  const sorted = Object.entries(DATA.aggregated).sort((a, b) => b[1][totalKey] - a[1][totalKey]);

  let rows = `<thead><tr>
    <th>#</th><th>Student</th><th>Classes</th><th>Prep avg</th><th>Part avg</th>
    <th>Att avg</th><th>Total / 20</th><th>% Component A</th>
  </tr></thead><tbody>`;

  sorted.forEach(([s, d], i) => {{
    const rankBadge = i < 3 ? `badge badge-rank-${{i + 1}}` : 'badge badge-rank-other';
    const pct = d[pctKey];
    rows += `<tr>
      <td><span class="${{rankBadge}}">${{i + 1}}°</span></td>
      <td class="student-name">${{s}}</td>
      <td>${{d.n_classes}}</td>
      <td>${{d.preparation_avg.toFixed(2)}}</td>
      <td>${{d[partKey].toFixed(2)}}</td>
      <td>${{d.attendance_avg.toFixed(2)}}</td>
      <td class="total">${{d[totalKey].toFixed(2)}}</td>
      <td>
        <span class="pct-bar"><div style="width:${{pct}}%"></div></span>
        ${{pct.toFixed(1)}}%
      </td>
    </tr>`;
  }});
  rows += '</tbody>';
  document.getElementById('rankingTable').innerHTML = rows;
}}

function renderEvolutionChart() {{
  const scenario = document.getElementById('scenarioSelect').value;
  const totalKey = scenario === 'A' ? 'total_a' : 'total_b';
  const datasets = Object.entries(DATA.aggregated).map(([s, d], i) => ({{
    label: s,
    data: d.by_class.map(bc => ({{ x: bc.session_id, y: bc[totalKey] }})),
    borderColor: COLORS[i],
    backgroundColor: COLORS[i] + '20',
    tension: 0.3,
    fill: false,
    pointRadius: 5,
    pointHoverRadius: 7,
  }}));

  if (evolutionChart) evolutionChart.destroy();
  evolutionChart = new Chart(document.getElementById('evolutionChart'), {{
    type: 'line',
    data: {{ datasets }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      scales: {{
        x: {{ type: 'category', labels: DATA.class_ids, title: {{ display: true, text: 'Class' }} }},
        y: {{ min: 0, max: 20, title: {{ display: true, text: 'Component A Total (out of 20)' }} }},
      }},
      plugins: {{
        legend: {{ position: 'top' }},
        tooltip: {{ mode: 'index', intersect: false }},
      }},
    }},
  }});
}}

function renderClassCards() {{
  const scenario = document.getElementById('scenarioSelect').value;
  const totalKey = scenario === 'A' ? 'component_a_total_scenario_A' : 'component_a_total_scenario_B';

  let html = '';
  CLASSES.forEach(c => {{
    const summary = c.summary;
    const top = Object.entries(summary).sort((a, b) => b[1][totalKey] - a[1][totalKey])[0];
    const groupAvg = Object.values(summary).reduce((s, v) => s + v[totalKey], 0) / Object.values(summary).length;

    html += `<a href="classes/${{c.session_id}}.html" class="class-card-link">
      <div class="class-card">
        <div class="date">${{c.session_id}}</div>
        <div class="meta">${{c.n_questions}} questions · ${{c.n_contributions}} contributions</div>
        <div class="stats-mini">
          <span>📊 Average: <strong>${{groupAvg.toFixed(2)}}</strong></span>
          <span>🏆 Top: <strong>${{top[0]}} (${{top[1][totalKey].toFixed(2)}})</strong></span>
        </div>
        <div class="card-arrow">View detail →</div>
      </div>
    </a>`;
  }});
  document.getElementById('classesGrid').innerHTML = html;
}}

let bRendered = false;
function safeRun(name, fn) {{
  try {{ fn(); }} catch (e) {{ console.error('[render ' + name + '] failed:', e); }}
}}
function renderAll() {{
  safeRun('stats', renderStats);
  safeRun('ranking', renderRanking);
  safeRun('evolution', renderEvolutionChart);
  safeRun('classCards', renderClassCards);
  safeRun('componentB', () => {{
    renderComponentBNoChart();  // tablas y stats (no chart hasta abrir tab)
  }});
  safeRun('componentC', renderComponentC);
}}

function renderComponentC() {{
  const C = COMPONENT_C;
  const studentsList = ['Aryang', 'Mega', 'Chilaka', 'Grace', 'Sthepen'];

  // Calcular totales por alumno (suma de todos los exámenes completados)
  const totals = {{}};
  studentsList.forEach(s => totals[s] = {{ sum: 0, max: 0, n_done: 0 }});
  C.exams.forEach(exam => {{
    if (exam.status !== 'completed' || !exam.responses) return;
    studentsList.forEach(s => {{
      const r = exam.responses[s];
      if (!r) return;
      const examMax = exam.n_questions * exam.max_per_question;
      const examScore = r.scores.reduce((a, b) => a + b, 0);
      totals[s].sum += examScore;
      totals[s].max += examMax;
      totals[s].n_done += 1;
    }});
  }});

  // ── C.1 Summary table ──
  const maxTotal = 5 * 4 * 25;  // 5 exams × 4 questions × 25 = 500
  let sumHtml = `<thead><tr>
    <th>Student</th>
    <th>Exams done</th>
    <th>Points obtained</th>
    <th>Max possible (so far)</th>
    <th>% of C.1 (10)</th>
    <th>Progress to 10%</th>
  </tr></thead><tbody>`;
  const sortedC = studentsList.map(s => {{
    const t = totals[s];
    return {{ s, ...t, pct: t.max ? (t.sum / maxTotal) * 10 : 0 }};
  }}).sort((a, b) => b.sum - a.sum);
  sortedC.forEach(({{ s, sum, max, n_done, pct }}) => {{
    const pctFinal = ((sum / maxTotal) * 10).toFixed(2);
    const pctProgress = ((sum / maxTotal) * 100).toFixed(1);
    sumHtml += `<tr>
      <td class="student-name">${{s}}</td>
      <td>${{n_done}} / 5</td>
      <td class="b-cell-done">${{sum}} / ${{maxTotal}}</td>
      <td>${{max || '—'}}</td>
      <td class="total">${{pctFinal}} / 10</td>
      <td><span class="pct-bar"><div style="width:${{pctProgress}}%"></div></span> ${{pctProgress}}%</td>
    </tr>`;
  }});
  sumHtml += '</tbody>';
  document.getElementById('c1SummaryTable').innerHTML = sumHtml;

  // ── Lista de exámenes ──
  let listHtml = '';
  C.exams.forEach((exam, idx) => {{
    const examMax = exam.n_questions ? exam.n_questions * exam.max_per_question : 100;
    const isCompleted = exam.status === 'completed';
    const statusBadge = isCompleted
      ? '<span class="status-pill status-rec">✅ Completed</span>'
      : '<span class="status-pill status-pending">⏳ Pending</span>';

    listHtml += `<div class="exam-card">
      <h3 class="exam-header collapsible-header" onclick="toggleSection('exam${{exam.id}}Content', 'exam${{exam.id}}Toggle')">
        <span><strong>${{exam.title}}</strong> · ${{exam.date_label}} ${{statusBadge}}</span>
        <span class="toggle-icon" id="exam${{exam.id}}Toggle">▶</span>
      </h3>
      <div id="exam${{exam.id}}Content" class="exam-content" style="display:none;">`;

    if (!isCompleted) {{
      listHtml += `<p style="padding:24px; text-align:center; color:#94a3b8; font-style:italic;">
        This exam has not been administered yet. Once it's done, scores will appear here.
      </p>`;
    }} else {{
      // Por cada estudiante, mostrar sus respuestas + scores
      studentsList.forEach(s => {{
        const r = exam.responses[s];
        if (!r) return;
        const totalScore = r.scores.reduce((a, b) => a + b, 0);
        listHtml += `<div class="student-block">
          <h4 class="student-header collapsible-header" onclick="toggleSection('exam${{exam.id}}-${{s}}-content', 'exam${{exam.id}}-${{s}}-toggle')">
            <span>👤 <strong>${{s}}</strong> <small style="color:#64748b">(${{r.full_name}})</small> — Total: <strong style="color:#305496">${{totalScore}}/${{examMax}}</strong></span>
            <span class="toggle-icon" id="exam${{exam.id}}-${{s}}-toggle">▶</span>
          </h4>
          <div id="exam${{exam.id}}-${{s}}-content" class="student-answers" style="display:none;">`;
        r.answers.forEach((ans, qi) => {{
          const q = exam.questions[qi];
          const score = r.scores[qi];
          listHtml += `<div class="answer-block">
            <div class="question-text"><strong>${{q}}</strong></div>
            <div class="answer-text">${{ans || '<em style="color:#94a3b8">(No answer)</em>'}}</div>
            <div class="answer-score">
              <label>Score:</label>
              <input type="number" min="0" max="${{exam.max_per_question}}" value="${{score}}" class="score-input" disabled>
              <span class="score-max">/ ${{exam.max_per_question}}</span>
            </div>
          </div>`;
        }});
        listHtml += `</div></div>`;
      }});
    }}
    listHtml += `</div></div>`;
  }});
  document.getElementById('c1ExamsList').innerHTML = listHtml;
}}

function renderComponentBNoChart() {{
  // Render tablas sin chart (chart se hace al abrir Tab B)
  renderComponentB(true);  // skipChart=true
}}

function renderComponentB(skipChart) {{
  const B = COMPONENT_B;
  const sessEl = document.getElementById('b1Session');
  if (sessEl) sessEl.textContent = B.session;

  // ── Component B Summary (B.1 obtained + B.2 pending) ──
  const studentsList = ['Aryang', 'Mega', 'Chilaka', 'Grace', 'Sthepen'];
  let sumHtml = `<thead><tr>
    <th>Student</th>
    <th>B.1 — Cross-eval<br><small>10% max</small></th>
    <th>B.2 — Prof. &amp; TA<br><small>15% max</small></th>
    <th>Total B<br><small>25% max</small></th>
    <th>% of B obtained so far</th>
  </tr></thead><tbody>`;
  const sortedBySum = studentsList.map(s => {{
    const d = B.summary[s];
    const b1 = d?.pct_b1 || 0;
    return {{ s, b1 }};
  }}).sort((a, b) => b.b1 - a.b1);

  sortedBySum.forEach(({{ s, b1 }}) => {{
    const totalSoFar = b1;
    const pctOfB = (totalSoFar / 25) * 100;
    sumHtml += `<tr>
      <td class="student-name">${{s}}</td>
      <td class="b-cell-done">${{b1.toFixed(2)}} / 10</td>
      <td class="b-cell-pending">— pending</td>
      <td class="total">${{totalSoFar.toFixed(2)}} / 25</td>
      <td>
        <span class="pct-bar"><div style="width:${{pctOfB}}%"></div></span>
        ${{pctOfB.toFixed(1)}}%
      </td>
    </tr>`;
  }});
  sumHtml += '</tbody>';
  document.getElementById('bSummaryTable').innerHTML = sumHtml;

  // ── B.2 Pending table (preparada para Prof. evaluation) ──
  const pendingDims = ['Clarity', 'Structure', 'Audience', 'Conciseness', 'Delivery'];
  let pendingHtml = `<thead><tr>
    <th>Presenter</th>
    ${{pendingDims.map(d => `<th>${{d}}</th>`).join('')}}
    <th>Extra notes</th>
    <th>Total / 30</th>
    <th>% of B.2 (15)</th>
  </tr></thead><tbody>`;
  studentsList.forEach(s => {{
    pendingHtml += `<tr>
      <td class="student-name">${{s}}</td>
      ${{pendingDims.map(() => `<td class="cell-empty">—</td>`).join('')}}
      <td class="cell-empty" style="font-style:italic;">—</td>
      <td class="cell-empty">— / 30</td>
      <td class="cell-empty">— / 15</td>
    </tr>`;
  }});
  pendingHtml += '</tbody>';
  document.getElementById('b2PendingTable').innerHTML = pendingHtml;

  // ── Stats cards ──
  const presenters = Object.entries(B.summary).filter(([s, d]) => d.n_evals > 0);
  const sortedB = [...presenters].sort((a, b) => b[1].avg_total - a[1].avg_total);
  const groupAvg = presenters.length ? presenters.reduce((s, [_, d]) => s + d.avg_total, 0) / presenters.length : 0;
  const totalEvals = B.evaluations.length;
  document.getElementById('b1Stats').innerHTML = `
    <div class="stat-card">
      <div class="label">Group average</div>
      <div class="value">${{groupAvg.toFixed(2)}}</div>
      <div class="sublabel">/ 30 points · ${{(groupAvg / 30 * 100).toFixed(1)}}%</div>
    </div>
    <div class="stat-card">
      <div class="label">Presenters evaluated</div>
      <div class="value">${{presenters.length}}</div>
      <div class="sublabel">${{totalEvals}} total cross-evaluations</div>
    </div>
    <div class="stat-card">
      <div class="label">Top presenter</div>
      <div class="value">${{sortedB[0]?.[0] || '—'}}</div>
      <div class="sublabel">${{sortedB[0]?.[1].avg_total.toFixed(2) || ''}} avg</div>
    </div>
    <div class="stat-card">
      <div class="label">Needs attention</div>
      <div class="value">${{sortedB[sortedB.length-1]?.[0] || '—'}}</div>
      <div class="sublabel">${{sortedB[sortedB.length-1]?.[1].avg_total.toFixed(2) || ''}} avg</div>
    </div>
  `;

  // ── Ranking ──
  let rows = `<thead><tr>
    <th>#</th><th>Presenter</th><th># Evals</th>
    <th>Clarity</th><th>Structure</th><th>Audience</th><th>Conciseness</th><th>Delivery</th>
    <th>Total / 30</th><th>% of B.1 (10)</th>
  </tr></thead><tbody>`;
  sortedB.forEach(([s, d], i) => {{
    const rankBadge = i < 3 ? `badge badge-rank-${{i + 1}}` : 'badge badge-rank-other';
    rows += `<tr>
      <td><span class="${{rankBadge}}">${{i + 1}}°</span></td>
      <td class="student-name">${{s}}</td>
      <td>${{d.n_evals}}</td>
      <td>${{d.by_dim.clarity.toFixed(2)}}</td>
      <td>${{d.by_dim.structure.toFixed(2)}}</td>
      <td>${{d.by_dim.audience.toFixed(2)}}</td>
      <td>${{d.by_dim.conciseness.toFixed(2)}}</td>
      <td>${{d.by_dim.delivery.toFixed(2)}}</td>
      <td class="total">${{d.avg_total.toFixed(2)}}</td>
      <td class="total">${{d.pct_b1.toFixed(2)}}</td>
    </tr>`;
  }});
  rows += '</tbody>';
  document.getElementById('b1RankingTable').innerHTML = rows;

  // ── Bar chart per dimension ──
  const dims = ['clarity', 'structure', 'audience', 'conciseness', 'delivery'];
  const dimsLabels = ['Clarity', 'Structure', 'Audience', 'Conciseness', 'Delivery'];
  const datasets = sortedB.map(([s, d], i) => ({{
    label: s,
    data: dims.map(dd => d.by_dim[dd]),
    backgroundColor: COLORS[i] + 'cc',
    borderColor: COLORS[i],
    borderWidth: 1,
  }}));
  if (!skipChart) {{
    if (b1DimChart) b1DimChart.destroy();
    const canvasB = document.getElementById('b1DimChart');
    if (canvasB) {{
      b1DimChart = new Chart(canvasB, {{
        type: 'bar',
        data: {{ labels: dimsLabels, datasets }},
        options: {{
          responsive: true, maintainAspectRatio: false,
          scales: {{
            y: {{ min: 0, max: 6, title: {{ display: true, text: 'Avg score (0–6)' }} }},
          }},
          plugins: {{ legend: {{ position: 'top' }} }},
        }},
      }});
    }}
  }}

  // ── Matrix evaluator × presenter ──
  const evaluators = Array.from(new Set(B.evaluations.map(e => e.evaluator))).sort();
  const presentersSet = Array.from(new Set(B.evaluations.map(e => e.presenter))).sort();
  let mat = '<thead><tr><th>Evaluator ↓ / Presenter →</th>';
  presentersSet.forEach(p => mat += `<th>${{p}}</th>`);
  mat += '</tr></thead><tbody>';
  evaluators.forEach(ev => {{
    mat += `<tr><td class="student-name">${{ev}}</td>`;
    presentersSet.forEach(p => {{
      const match = B.evaluations.find(e => e.evaluator === ev && e.presenter === p);
      if (match) {{
        const tooltip = `Clarity ${{match.clarity}} · Structure ${{match.structure}} · Audience ${{match.audience}} · Conciseness ${{match.conciseness}} · Delivery ${{match.delivery}}`;
        const cls = match.total >= 28 ? 'cell-high' : match.total >= 22 ? 'cell-mid' : 'cell-low';
        mat += `<td class="cell ${{cls}}" title="${{tooltip}}">${{match.total}}</td>`;
      }} else if (ev === p) {{
        mat += `<td class="cell-self">—</td>`;
      }} else {{
        mat += `<td class="cell-empty">—</td>`;
      }}
    }});
    mat += '</tr>';
  }});
  mat += '</tbody>';
  document.getElementById('b1Matrix').innerHTML = mat;

  // ── Comments table ──
  let comms = `<thead><tr><th>Time</th><th>Evaluator</th><th>Presenter</th><th>Score</th><th>Comment</th></tr></thead><tbody>`;
  B.evaluations.filter(e => e.comment && e.comment.trim()).forEach(e => {{
    comms += `<tr>
      <td class="ts">${{e.timestamp}}</td>
      <td>${{e.evaluator}}</td>
      <td class="student-name">${{e.presenter}}</td>
      <td class="total">${{e.total}}</td>
      <td>${{e.comment}}</td>
    </tr>`;
  }});
  comms += '</tbody>';
  document.getElementById('b1Comments').innerHTML = comms;
}}

renderAll();
</script>
</body>
</html>
"""
    out_file = OUT_DIR / "dashboard.html"
    with open(out_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ Dashboard HTML: {out_file.relative_to(REPO_ROOT)}")


def main():
    print("=" * 70)
    print("  AGREGADOR — Componente A acumulado")
    print("=" * 70)

    classes = load_all_classes()
    if not classes:
        print("❌ No hay clases procesadas. Corre process_class.py primero.")
        sys.exit(1)

    print(f"\n📂 Clases encontradas: {len(classes)}")
    for c in classes:
        print(f"   · {c['session_id']}")

    per_student = aggregate(classes)

    print("\n📊 ACUMULADO POR ALUMNO (Escenario A):\n")
    sorted_students = sorted(per_student.items(), key=lambda x: x[1]["total_a_avg"], reverse=True)
    print(f"  {'#':>2}  {'Alumno':10s}  {'Clases':>6s}  {'Prep':>6s}  {'Part A':>6s}  {'Att':>5s}  {'Total A':>8s}  {'% A':>5s}")
    for i, (s, d) in enumerate(sorted_students, start=1):
        print(f"  {i:>2}  {s:10s}  {d['n_classes']:>6d}  {d['preparation_avg']:>6.2f}  "
              f"{d['participation_a_avg']:>6.2f}  {d['attendance_avg']:>5.2f}  "
              f"{d['total_a_avg']:>8.2f}  {d['pct_a_avg']:>4.1f}%")

    agg_data = write_json(per_student, classes)
    write_excel(per_student, classes)
    write_dashboard_html(per_student, classes, agg_data)

    print(f"\n✅ Listo. Para ver el dashboard:")
    print(f"   xdg-open {OUT_DIR.relative_to(REPO_ROOT)}/dashboard.html")
    print(f"   o copia ese HTML a tu Windows y ábrelo en el navegador")


if __name__ == "__main__":
    main()
